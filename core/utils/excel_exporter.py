"""
Excel Export utility
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export data to Excel with professional formatting."""
    
    def __init__(self):
        self.buffer = BytesIO()
    
    def export_report(self, data):
        """Export report data to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        # Title
        ws['A1'] = data.get('type', 'Rapport')
        ws['A1'].font = Font(size=16, bold=True, color="1E3A8A")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Period
        ws['A2'] = f"Période: {data.get('period', 'N/A')}"
        ws['A2'].font = Font(size=11, bold=True)
        ws.merge_cells('A2:D2')
        
        # Generate content based on report type
        report_type = data.get('type', '')
        
        if 'Compte de Résultat' in report_type or 'Profit' in report_type:
            self._export_profit_loss(ws, data)
        elif 'Vente' in report_type or 'Sales' in report_type:
            self._export_sales(ws, data)
        elif 'Dépenses' in report_type or 'Expenses' in report_type:
            self._export_expenses(ws, data)
        elif 'Inventaire' in report_type or 'Inventory' in report_type:
            self._export_inventory(ws, data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.buffer)
        self.buffer.seek(0)
        return self.buffer
    
    def _export_profit_loss(self, ws, data):
        """Export profit & loss data."""
        row = 4
        
        # Revenue section
        ws[f'A{row}'] = 'REVENUS'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1E40AF")
        row += 1
        
        # Headers
        ws[f'A{row}'] = 'Description'
        ws[f'B{row}'] = 'Montant (FCFA)'
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        revenue = data.get('revenue', {})
        ws[f'A{row}'] = 'Ventes'
        ws[f'B{row}'] = revenue.get('sales', 0)
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Factures'
        ws[f'B{row}'] = revenue.get('invoices', 0)
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
        
        ws[f'A{row}'] = 'Total Revenus'
        ws[f'B{row}'] = revenue.get('total', 0)
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        ws[f'B{row}'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = 'DÉPENSES'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="DC2626")
        row += 1
        
        ws[f'A{row}'] = 'Description'
        ws[f'B{row}'] = 'Montant (FCFA)'
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        ws[f'A{row}'] = 'Total Dépenses'
        ws[f'B{row}'] = data.get('expenses', 0)
        ws[f'B{row}'].number_format = '#,##0'
        row += 2
        
        # Profit section
        ws[f'A{row}'] = 'RÉSULTAT'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="64748B")
        row += 1
        
        ws[f'A{row}'] = 'Description'
        ws[f'B{row}'] = 'Montant (FCFA)'
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        profit = data.get('profit', 0)
        ws[f'A{row}'] = 'Bénéfice Net' if profit >= 0 else 'Perte Nette'
        ws[f'B{row}'] = profit
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        color = "10B981" if profit >= 0 else "DC2626"
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _export_sales(self, ws, data):
        """Export sales data."""
        row = 4
        
        # Headers
        headers = ['Date', 'Montant (FCFA)', 'Vendeur', 'Mode Paiement']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data
        sales = data.get('sales', [])
        total = 0
        for sale in sales:
            ws.cell(row=row, column=1, value=str(sale.get('sale_date', '')))
            ws.cell(row=row, column=2, value=sale.get('total_amount', 0))
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3, value=sale.get('user__username', 'N/A'))
            ws.cell(row=row, column=4, value=sale.get('payment_method', 'N/A'))
            total += sale.get('total_amount', 0)
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=2).number_format = '#,##0'
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    def _export_expenses(self, ws, data):
        """Export expenses data."""
        row = 4
        
        # Headers
        headers = ['Date', 'Montant (FCFA)', 'Catégorie', 'Description']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data
        expenses = data.get('expenses', [])
        total = 0
        for expense in expenses:
            ws.cell(row=row, column=1, value=str(expense.get('expense_date', '')))
            ws.cell(row=row, column=2, value=expense.get('amount', 0))
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3, value=expense.get('category__name', 'N/A'))
            ws.cell(row=row, column=4, value=expense.get('description', ''))
            total += expense.get('amount', 0)
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=2).number_format = '#,##0'
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    
    def _export_inventory(self, ws, data):
        """Export inventory data."""
        row = 4
        
        # Headers
        headers = ['Produit', 'Quantité', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data
        stocks = data.get('stocks', [])
        total_value = 0
        for stock in stocks:
            quantity = stock.get('quantity', 0)
            unit_price = stock.get('unit_price', 0)
            value = quantity * unit_price
            
            ws.cell(row=row, column=1, value=stock.get('product__name', 'N/A'))
            ws.cell(row=row, column=2, value=quantity)
            ws.cell(row=row, column=3, value=unit_price)
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4, value=value)
            ws.cell(row=row, column=4).number_format = '#,##0'
            total_value += value
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=4, value=total_value)
        ws.cell(row=row, column=4).number_format = '#,##0'
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
