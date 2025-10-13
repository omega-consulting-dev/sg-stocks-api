import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas


class ExcelExporter:
    """Utility class for Excel exports."""
    
    @staticmethod
    def create_workbook(title="Export"):
        """Create a new workbook with basic styling."""
        wb = Workbook()
        ws = wb.active
        ws.title = title
        return wb, ws
    
    @staticmethod
    def style_header(ws, columns, row=1):
        """Apply header styling to the first row."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
    
    @staticmethod
    def auto_adjust_columns(ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def generate_response(wb, filename):
        """Generate HTTP response with Excel file."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PDFExporter:
    """Utility class for PDF exports."""
    
    @staticmethod
    def create_document(buffer, pagesize=A4):
        """Create a PDF document."""
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        return doc
    
    @staticmethod
    def get_styles():
        """Get custom paragraph styles."""
        styles = getSampleStyleSheet()
        
        # Title style
        styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        ))
        
        # Subtitle style
        styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#666666'),
            spaceAfter=20,
        ))
        
        return styles
    
    @staticmethod
    def create_table(data, col_widths=None):
        """Create a styled table."""
        table = Table(data, colWidths=col_widths)
        
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        table.setStyle(style)
        return table
    
    @staticmethod
    def generate_response(buffer, filename):
        """Generate HTTP response with PDF file."""
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ImportValidator:
    """Utility class for validating imports."""
    
    @staticmethod
    def validate_required_fields(row, required_fields):
        """Validate that all required fields are present."""
        errors = []
        for field in required_fields:
            if field not in row or not row[field]:
                errors.append(f"Le champ '{field}' est obligatoire")
        return errors
    
    @staticmethod
    def validate_unique_field(model, field_name, value, exclude_id=None):
        """Validate that a field value is unique."""
        query = model.objects.filter(**{field_name: value})
        if exclude_id:
            query = query.exclude(id=exclude_id)
        return not query.exists()
    
    @staticmethod
    def validate_foreign_key(model, field_name, value):
        """Validate that a foreign key exists."""
        try:
            return model.objects.get(**{field_name: value})
        except model.DoesNotExist:
            return None

