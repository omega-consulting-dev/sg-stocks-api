"""
Inventory views for API.
"""

from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import HasModulePermission
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view
from django.db.models import Sum, F, Q
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
import django_filters

from core.utils.export_utils import ExcelExporter, PDFExporter
from core.mixins import StoreAccessMixin, PermissionCheckMixin, UserStoreValidationMixin
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib.units import inch
import io

from apps.inventory.models import (
    Store, Stock, StockMovement, StockTransfer, 
    StockTransferLine, Inventory, InventoryLine, ReceiptNumberSequence
)
from apps.inventory.serializers import (
    StoreSerializer, StockSerializer, StockMovementSerializer,
    StockTransferListSerializer, StockTransferDetailSerializer,
    StockTransferCreateSerializer, InventoryListSerializer,
    InventoryDetailSerializer, InventoryCreateSerializer
)


@extend_schema_view(
    list=extend_schema(summary="Liste des magasins", tags=["Inventory"]),
    retrieve=extend_schema(summary="Détail d'un magasin", tags=["Inventory"]),
    create=extend_schema(summary="Créer un magasin", tags=["Inventory"]),
    update=extend_schema(summary="Modifier un magasin", tags=["Inventory"]),
)
class StoreViewSet(viewsets.ModelViewSet):
    """ViewSet for Store model."""
    
    queryset = Store.objects.filter(is_active=True).select_related('manager')
    serializer_class = StoreSerializer
    permission_classes = [IsAuthenticated]  # Pas de HasModulePermission - tous peuvent voir leurs stores assignés
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['store_type', 'is_active']
    search_fields = ['name', 'code', 'city']
    ordering_fields = ['name', 'code', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope in ['assigned', 'own']:
                # Les utilisateurs voient leurs stores assignés
                return queryset.filter(id__in=user.assigned_stores.values_list('id', flat=True))
        
        # Par défaut, retourner les stores assignés
        return queryset.filter(id__in=user.assigned_stores.values_list('id', flat=True))
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
    
    def perform_destroy(self, instance):
        """Soft delete: désactiver au lieu de supprimer."""
        instance.is_active = False
        instance.updated_by = self.request.user
        instance.save()
    
    @extend_schema(summary="Statistiques d'un magasin", tags=["Inventory"])
    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Get store statistics."""
        store = self.get_object()
        stats = {
            'total_products': store.stocks.count(),
            'total_stock_value': store.stocks.aggregate(
                total=Sum(F('quantity') * F('product__cost_price'))
            )['total'] or 0,
            'low_stock_items': store.stocks.filter(
                quantity__lt=F('product__minimum_stock')
            ).count(),
        }
        return Response(stats)
    
    @extend_schema(
        summary="Liste des produits d'un magasin",
        description="Récupère tous les produits présents dans un magasin avec leur stock.",
        tags=["Inventory"]
    )
    @action(detail=True, methods=['get'])
    def products(self, request, pk=None):
        """Get all products in a store with their stock information."""
        store = self.get_object()
        
        # Récupérer tous les stocks de ce magasin avec les infos produits
        stocks = Stock.objects.filter(store=store).select_related(
            'product', 'product__category'
        ).order_by('product__name')
        
        # Préparer les données
        products_data = []
        for stock in stocks:
            product = stock.product
            products_data.append({
                'product_id': product.id,
                'product_name': product.name,
                'product_reference': product.reference,
                'category': product.category.name if product.category else None,
                'quantity': float(stock.quantity),
                'reserved_quantity': float(stock.reserved_quantity),
                'available_quantity': float(stock.available_quantity),
                'minimum_stock': product.minimum_stock,
                'optimal_stock': product.optimal_stock,
                'cost_price': float(product.cost_price) if product.cost_price else None,
                'selling_price': float(product.selling_price) if product.selling_price else None,
                'stock_status': 'low' if stock.quantity < product.minimum_stock else 'normal',
                'is_active': product.is_active,
            })
        
        return Response({
            'store': {
                'id': store.id,
                'name': store.name,
                'code': store.code,
            },
            'total_products': len(products_data),
            'products': products_data
        })


@extend_schema_view(
    list=extend_schema(summary="Liste des stocks", tags=["Inventory"]),
    retrieve=extend_schema(summary="Détail d'un stock", tags=["Inventory"]),
)
class StockViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Stock model (read-only)."""
    
    queryset = Stock.objects.select_related('product', 'store')
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated]  # Lecture seule, tous les utilisateurs authentifiés peuvent voir les stocks
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['product', 'store']
    search_fields = ['product__name', 'product__reference']
    ordering_fields = ['quantity', 'created_at']
    ordering = ['created_at']
    
    def list(self, request, *args, **kwargs):
        """Override list to support historical stock calculation."""
        date_to = request.query_params.get('date_to')
        date_from = request.query_params.get('date_from')
        
        if date_to or date_from:
            # Calculate historical stock for the specified period
            return self.calculate_historical_stocks(date_from, date_to, request)
        
        # Default behavior for current stocks
        return super().list(request, *args, **kwargs)
    
    def calculate_historical_stocks(self, date_from, date_to, request):
        """Calculate stock quantities for a specific period."""
        from datetime import datetime
        from django.db.models import Sum, Q
        from decimal import Decimal
        from django.utils import timezone as django_timezone
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f"Calculating historical stocks - from: {date_from}, to: {date_to}")
        
        # Parse and validate dates
        target_date_start = None
        target_date_end = None
        
        try:
            if date_from:
                target_date_start = datetime.strptime(date_from, '%Y-%m-%d')
                target_date_start = target_date_start.replace(hour=0, minute=0, second=0)
                target_date_start = django_timezone.make_aware(target_date_start, django_timezone.get_current_timezone())
                logger.info(f"Start date parsed: {target_date_start}")
            
            if date_to:
                target_date_end = datetime.strptime(date_to, '%Y-%m-%d')
                target_date_end = target_date_end.replace(hour=23, minute=59, second=59)
                target_date_end = django_timezone.make_aware(target_date_end, django_timezone.get_current_timezone())
                logger.info(f"End date parsed: {target_date_end}")
        except ValueError as e:
            logger.error(f"Invalid date format, error: {e}")
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD'},
                status=400
            )
        
        # Get all product-store combinations
        stocks = Stock.objects.select_related('product', 'store').all()
        
        # Apply filters
        product_filter = request.query_params.get('product')
        store_filter = request.query_params.get('store')
        search_filter = request.query_params.get('search')
        
        if product_filter:
            stocks = stocks.filter(product_id=product_filter)
            logger.info(f"Filtering by product: {product_filter}")
        if store_filter:
            stocks = stocks.filter(store_id=store_filter)
            logger.info(f"Filtering by store: {store_filter}")
        if search_filter:
            stocks = stocks.filter(
                Q(product__name__icontains=search_filter) | 
                Q(product__reference__icontains=search_filter)
            )
            logger.info(f"Filtering by search: {search_filter}")
        
        logger.info(f"Total stocks to process: {stocks.count()}")
        historical_stocks = []
        
        for stock in stocks:
            # Build movement filter - calcul jusqu'à la date de fin (tous les mouvements avant date_to)
            # On utilise le champ 'date' (date de réalisation) au lieu de 'created_at'
            movement_filter = Q(
                product=stock.product,
                store=stock.store,
                is_active=True
            )
            
            # Filtrer jusqu'à la date de fin si spécifiée
            # On veut tous les mouvements AVANT et JUSQU'AU date_to pour avoir le stock à cette date
            if target_date_end:
                movement_filter &= Q(date__lte=target_date_end.date())
            
            # Optionnellement, commencer à partir d'une date de début
            # Ceci permet de voir le stock à une période précise si besoin
            # if target_date_start:
            #     movement_filter &= Q(date__gte=target_date_start.date())
            
            # Calculate movements jusqu'à la date
            movements = StockMovement.objects.filter(movement_filter)
            
            # Sum entries
            entrees = movements.filter(movement_type='in').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Sum exits
            sorties = movements.filter(movement_type='out').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Sum transfer outs
            transferts_out = movements.filter(movement_type='transfer').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Sum transfer ins jusqu'à la date
            transfer_in_filter = Q(
                product=stock.product,
                destination_store=stock.store,
                movement_type='transfer',
                is_active=True
            )
            if target_date_end:
                transfer_in_filter &= Q(date__lte=target_date_end.date())
                
            transferts_in = StockMovement.objects.filter(transfer_in_filter).aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Calculate historical quantity
            historical_quantity = entrees - sorties - transferts_out + transferts_in
            
            # Create a copy of stock with historical quantity
            historical_stock = {
                'id': stock.id,
                'product': stock.product.id,
                'product_name': stock.product.name,
                'product_reference': stock.product.reference or '',
                'store': stock.store.id,
                'store_name': stock.store.name,
                'quantity': float(historical_quantity),
                'reserved_quantity': 0,  # No reserved quantity for historical data
                'available_quantity': float(historical_quantity),
                'created_at': stock.created_at.isoformat(),
                'updated_at': stock.updated_at.isoformat(),
            }
            historical_stocks.append(historical_stock)
        
        logger.info(f"Historical stocks calculated: {len(historical_stocks)} items")
        
        # Pagination
        page = self.paginate_queryset(historical_stocks)
        if page is not None:
            return self.get_paginated_response(page)
        
        return Response(historical_stocks)
    
    @extend_schema(summary="Produits en rupture", tags=["Inventory"])
    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Get low stock items."""
        queryset = self.get_queryset().filter(
            quantity__lt=F('product__minimum_stock')
        )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export stock levels to Excel with filtering."""
        from datetime import datetime
        from django.db.models import Sum, Q
        from decimal import Decimal
        from django.utils import timezone as django_timezone
        import logging
        
        logger = logging.getLogger(__name__)
        
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        if date_to or date_from:
            # Calculate historical stocks (sans pagination pour l'export)
            target_date_start = None
            target_date_end = None
            
            try:
                if date_from:
                    target_date_start = datetime.strptime(date_from, '%Y-%m-%d')
                    target_date_start = target_date_start.replace(hour=0, minute=0, second=0)
                    target_date_start = django_timezone.make_aware(target_date_start, django_timezone.get_current_timezone())
                
                if date_to:
                    target_date_end = datetime.strptime(date_to, '%Y-%m-%d')
                    target_date_end = target_date_end.replace(hour=23, minute=59, second=59)
                    target_date_end = django_timezone.make_aware(target_date_end, django_timezone.get_current_timezone())
            except ValueError:
                pass
            
            # Get all product-store combinations
            stocks = Stock.objects.select_related('product', 'store').all()
            
            # Apply filters
            product_filter = request.query_params.get('product')
            store_filter = request.query_params.get('store')
            search_filter = request.query_params.get('search')
            
            if product_filter:
                stocks = stocks.filter(product_id=product_filter)
            if store_filter:
                stocks = stocks.filter(store_id=store_filter)
            if search_filter:
                stocks = stocks.filter(
                    Q(product__name__icontains=search_filter) | 
                    Q(product__reference__icontains=search_filter)
                )
            
            stocks_list = []
            for stock in stocks:
                # Calculer le stock historique
                movement_filter = Q(
                    product=stock.product,
                    store=stock.store,
                    is_active=True
                )
                
                if target_date_end:
                    movement_filter &= Q(date__lte=target_date_end.date())
                
                movements = StockMovement.objects.filter(movement_filter)
                
                entrees = movements.filter(movement_type='in').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                sorties = movements.filter(movement_type='out').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                transferts_out = movements.filter(movement_type='transfer').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                transfer_in_filter = Q(
                    product=stock.product,
                    destination_store=stock.store,
                    movement_type='transfer',
                    is_active=True
                )
                if target_date_end:
                    transfer_in_filter &= Q(date__lte=target_date_end.date())
                    
                transferts_in = StockMovement.objects.filter(transfer_in_filter).aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                historical_quantity = entrees - sorties - transferts_out + transferts_in
                
                stocks_list.append({
                    'product_name': stock.product.name,
                    'product_reference': stock.product.reference or '',
                    'store_name': stock.store.name,
                    'quantity': float(historical_quantity),
                    'minimum_stock': stock.product.minimum_stock or 0
                })
        else:
            # Use current stocks with filters
            stocks = self.filter_queryset(self.get_queryset())
            stocks_list = stocks
        
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventaire"
        
        # Colonnes comme dans le tableau frontend
        columns = [
            'Référence', 'Produit / Désignation', 'Magasin', 'Stock théorique', 'Stock physique', 'Ecart'
        ]
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, col_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_num, stock in enumerate(stocks_list, 2):
            # Handle both Stock objects and dictionaries (historical stocks)
            if isinstance(stock, dict):
                product_name = stock['product_name']
                product_reference = stock.get('product_reference', '') or '-'
                store_name = stock['store_name']
                quantity = stock['quantity']
            else:
                product_name = stock.product.name
                product_reference = stock.product.reference or '-'
                store_name = stock.store.name
                quantity = float(stock.quantity)
            
            # Colonnes du tableau:
            # 1. Référence
            ws.cell(row=row_num, column=1, value=product_reference)
            # 2. Produit / Désignation
            ws.cell(row=row_num, column=2, value=product_name)
            # 3. Magasin
            ws.cell(row=row_num, column=3, value=store_name)
            # 4. Stock théorique (le stock calculé/actuel)
            ws.cell(row=row_num, column=4, value=quantity)
            # 5. Stock physique (vide - pour saisie manuelle)
            ws.cell(row=row_num, column=5, value='')
            # 6. Écart (vide - pour saisie manuelle)
            ws.cell(row=row_num, column=6, value='')
        
        # Auto adjust columns
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Generate response
        response = HttpResponse(content_type='application/vnd.ms-excel')
        date_suffix = f"_au_{date_to.replace('-', '')}" if date_to else ""
        filename = f"inventaire{date_suffix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response


class StockMovementFilterSet(django_filters.FilterSet):
    """Custom FilterSet for StockMovement with date range filters."""
    date_from = django_filters.DateFilter(field_name='date', lookup_expr='gte')
    date_to = django_filters.DateFilter(field_name='date', lookup_expr='lte')
    movement_type = django_filters.BaseInFilter(field_name='movement_type', lookup_expr='in')
    
    class Meta:
        model = StockMovement
        fields = ['product', 'store', 'movement_type', 'supplier', 'date_from', 'date_to']


@extend_schema_view(
    list=extend_schema(summary="Liste des mouvements", tags=["Inventory"]),
    retrieve=extend_schema(summary="Détail d'un mouvement", tags=["Inventory"]),
    create=extend_schema(summary="Créer un mouvement", tags=["Inventory"]),
)
class StockMovementViewSet(StoreAccessMixin, UserStoreValidationMixin, viewsets.ModelViewSet):
    """ViewSet for StockMovement model with automatic store filtering."""
    
    queryset = StockMovement.objects.filter(is_active=True).select_related(
        'product', 'store', 'destination_store','supplier', 'created_by', 'purchase_order'
    ).prefetch_related('purchase_order__payments')
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'inventory'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = StockMovementFilterSet
    search_fields = ['product__name', 'reference', 'supplier__name', 'supplier__supplier_code']
    ordering_fields = ['date', 'created_at']
    ordering = ['date', 'created_at']  # Tri par date de réalisation (du plus ancien au plus récent)
    http_method_names = ['get', 'post', 'put', 'patch', 'delete', 'head', 'options']

    def get_queryset(self):
        """
        Override pour gérer les mouvements visibles dans les deux magasins (source et destination).
        Pour les transferts, un mouvement doit être visible si l'utilisateur a accès
        soit au magasin source, soit au magasin destination.
        """
        queryset = super().get_queryset()
        user = self.request.user
        
        # Super admin voit tout
        if user.is_superuser:
            return queryset
        
        # Vérifier si l'utilisateur a un rôle
        if not hasattr(user, 'role') or not user.role:
            return queryset.none()
        
        # Si le rôle a accès à tous les stores
        if user.role.access_scope == 'all':
            return queryset
        
        # Si le rôle a accès uniquement aux stores assignés
        if user.role.access_scope == 'assigned':
            # Pour les mouvements, on doit voir ceux où le store OU le destination_store
            # est dans les stores assignés (pour inclure les transferts entrants et sortants)
            from django.db.models import Q
            return queryset.filter(
                Q(store__in=user.assigned_stores.all()) | 
                Q(destination_store__in=user.assigned_stores.all())
            )
        
        # Si le rôle a accès uniquement à ses propres données
        if user.role.access_scope == 'own':
            return queryset.filter(created_by=user)
        
        return queryset

    @extend_schema(summary="Obtenir le prochain numéro de pièce", tags=["Inventory"])
    @action(detail=False, methods=['get'], url_path='next-receipt-number', permission_classes=[IsAuthenticated])
    def next_receipt_number(self, request):
        """Get the next available receipt number."""
        # Récupérer le plus grand numéro existant (actif ou non) pour continuer la séquence
        all_receipts = StockMovement.objects.filter(
            receipt_number__startswith='RECEIPT-'
        ).values_list('receipt_number', flat=True)
        
        numbers = []
        for receipt in all_receipts:
            try:
                num = int(receipt.replace('RECEIPT-', ''))
                numbers.append(num)
            except (ValueError, AttributeError):
                pass
        
        next_num = max(numbers) + 1 if numbers else 1
        next_receipt = f'RECEIPT-{str(next_num).zfill(3)}'
        return Response({'next_receipt_number': next_receipt})

    
    def perform_create(self, serializer):
        """Create a stock movement and update stock."""
        movement = serializer.save(created_by=self.request.user)
        self._update_stock(movement)
    
    def perform_update(self, serializer):
        """Override to handle stock updates when modifying a movement."""
        from django.db import transaction
        
        # Utiliser une transaction atomique pour garantir la cohérence
        with transaction.atomic():
            old_instance = self.get_object()
            
            # Inverser l'ancien mouvement
            self._reverse_stock(old_instance)
            
            # Sauvegarder les modifications
            updated_instance = serializer.save(updated_by=self.request.user)
            
            # Appliquer le nouveau mouvement
            self._update_stock(updated_instance)
    
    def perform_destroy(self, instance):
        """Override to reverse stock changes and soft delete."""
        from django.db import transaction
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f"[DELETE] Tentative de suppression du mouvement {instance.id} - {instance.movement_type} - "
                   f"{instance.product.name} - Quantité: {instance.quantity}")
        
        # Utiliser une transaction atomique pour garantir la cohérence
        try:
            with transaction.atomic():
                # Inverser les changements de stock avant la désactivation
                self._reverse_stock(instance)
                
                # Soft delete: désactiver au lieu de supprimer
                instance.is_active = False
                instance.updated_by = self.request.user
                instance.save()
                
                logger.info(f"[DELETE] ✓ Mouvement {instance.id} supprimé avec succès et stock inversé")
        except ValidationError as e:
            logger.error(f"[DELETE] ✗ Erreur lors de la suppression du mouvement {instance.id}: {e}")
            raise
        except Exception as e:
            logger.error(f"[DELETE] ✗ Erreur inattendue lors de la suppression du mouvement {instance.id}: {e}")
            raise ValidationError(f"Erreur lors de la suppression: {str(e)}")
    
    def _reverse_stock(self, movement):
        """Reverse stock changes when deleting a movement."""
        from django.db import transaction
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Utiliser select_for_update pour éviter les problèmes de concurrence
        with transaction.atomic():
            try:
                # Verrouiller la ligne pour éviter les conditions de course
                stock = Stock.objects.select_for_update().get(
                    product=movement.product,
                    store=movement.store
                )
                
                logger.info(f"[REVERSE] Stock avant annulation: {stock.quantity} pour {movement.product.name}")
                
            except Stock.DoesNotExist:
                # Si le stock n'existe pas, on ne peut pas inverser
                logger.error(f"[REVERSE] Stock introuvable pour {movement.product.name} dans {movement.store.name}")
                raise ValidationError(
                    f"Impossible de supprimer ce mouvement. Le stock pour {movement.product.name} "
                    f"dans {movement.store.name} n'existe plus."
                )
            
            if movement.movement_type == 'in':
                # Inverser une entrée: soustraire la quantité
                if stock.quantity < movement.quantity:
                    logger.error(f"[REVERSE] Stock insuffisant ({stock.quantity}) pour annuler l'entrée ({movement.quantity})")
                    raise ValidationError(
                        f"Impossible de supprimer cette entrée. Le stock actuel de {movement.product.name} "
                        f"({stock.quantity}) est inférieur à la quantité de l'entrée ({movement.quantity}). "
                        "Des sorties ont probablement été effectuées depuis cette entrée. "
                        "Veuillez d'abord annuler ces sorties."
                    )
                stock.quantity -= movement.quantity
                logger.info(f"[REVERSE] Entrée annulée: {stock.quantity + movement.quantity} → {stock.quantity}")
                stock.save(update_fields=['quantity', 'updated_at'])
                
            elif movement.movement_type == 'out':
                # Inverser une sortie: ajouter la quantité
                stock.quantity += movement.quantity
                logger.info(f"[REVERSE] Sortie annulée: {stock.quantity - movement.quantity} → {stock.quantity}")
                stock.save(update_fields=['quantity', 'updated_at'])
                
            elif movement.movement_type == 'transfer':
                # Inverser un transfert: rajouter au stock source et retirer de la destination
                stock.quantity += movement.quantity
                logger.info(f"[REVERSE] Transfert annulé (source): {stock.quantity - movement.quantity} → {stock.quantity}")
                stock.save(update_fields=['quantity', 'updated_at'])
                
                if movement.destination_store:
                    try:
                        dest_stock = Stock.objects.select_for_update().get(
                            product=movement.product,
                            store=movement.destination_store
                        )
                        if dest_stock.quantity < movement.quantity:
                            logger.error(f"[REVERSE] Stock destination insuffisant ({dest_stock.quantity}) pour annuler le transfert ({movement.quantity})")
                            raise ValidationError(
                                f"Impossible de supprimer ce transfert. Le stock de {movement.product.name} "
                                f"au magasin destination ({dest_stock.quantity}) est inférieur à la quantité "
                                f"transférée ({movement.quantity})."
                            )
                        dest_stock.quantity -= movement.quantity
                        logger.info(f"[REVERSE] Transfert annulé (destination): {dest_stock.quantity + movement.quantity} → {dest_stock.quantity}")
                        dest_stock.save(update_fields=['quantity', 'updated_at'])
                    except Stock.DoesNotExist:
                        # Si le stock destination n'existe pas, c'est ok pour la suppression
                        logger.warning(f"[REVERSE] Stock destination introuvable, annulation partielle du transfert")
                        pass
     
    def _update_stock(self, movement):
        """Update stock based on movement type."""
        from django.db import transaction
        
        # Utiliser une transaction atomique pour garantir la cohérence
        with transaction.atomic():
            # Utiliser select_for_update pour éviter les conditions de course
            stock, created = Stock.objects.select_for_update().get_or_create(
                product=movement.product,
                store=movement.store,
                defaults={'quantity': 0, 'reserved_quantity': 0}
            )
            
            if movement.movement_type == 'in':
                stock.quantity += movement.quantity
                
            elif movement.movement_type == 'out':
                # Vérification supplémentaire (normalement déjà validé par le serializer)
                if stock.quantity < movement.quantity:
                    raise ValidationError(
                        f"Stock insuffisant pour {movement.product.name} dans {movement.store.name}. "
                        f"Disponible: {stock.quantity}, Demandé: {movement.quantity}"
                    )
                stock.quantity -= movement.quantity
                
            elif movement.movement_type == 'transfer':
                # Pour les transferts, diminuer le stock source
                if stock.quantity < movement.quantity:
                    raise ValidationError(
                        f"Stock insuffisant pour transférer {movement.product.name}. "
                        f"Disponible: {stock.quantity}, Demandé: {movement.quantity}"
                    )
                stock.quantity -= movement.quantity
                
                # Augmenter le stock destination si spécifié
                if movement.destination_store:
                    dest_stock, _ = Stock.objects.select_for_update().get_or_create(
                        product=movement.product,
                        store=movement.destination_store,
                        defaults={'quantity': 0, 'reserved_quantity': 0}
                    )
                    dest_stock.quantity += movement.quantity
                    dest_stock.save(update_fields=['quantity', 'updated_at'])
            
            stock.save(update_fields=['quantity', 'updated_at'])

    @extend_schema(summary="Exporter les mouvements en Excel", tags=["Inventory"])
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export stock movements to Excel with filtering.
        Supports all filters from the list endpoint:
        - movement_type: 'in', 'out', 'transfer'
        - date_from: YYYY-MM-DD (applies to created_at)
        - date_to: YYYY-MM-DD (applies to created_at)
        - supplier: supplier id
        - product: product id
        - store: store id
        - search: search in product name, reference, supplier name, supplier code
        """
        # Use filter_queryset to apply all configured filters (DjangoFilterBackend, SearchFilter, etc.)
        movements = self.filter_queryset(self.get_queryset())
        
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mouvements"
        
        columns = [
            'Date', 'Désignation', 'Entrées', 'Sorties', 'Stocks final'
        ]
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, col_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sort movements by date (oldest first) for progressive stock calculation
        movements_list = list(movements.order_by('created_at'))
        
        # Calculate progressive stock
        stock_cumulatif = 0
        
        # Add data
        for row_num, movement in enumerate(movements_list, 2):
            entrees = ''
            sorties = ''
            
            if movement.movement_type == 'in':
                entrees = float(movement.quantity)
                stock_cumulatif += entrees
            elif movement.movement_type == 'out':
                sorties = float(movement.quantity)
                stock_cumulatif -= sorties
            
            ws.cell(row=row_num, column=1, value=movement.created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=2, value=movement.product.name)
            ws.cell(row=row_num, column=3, value=entrees)
            ws.cell(row=row_num, column=4, value=sorties)
            ws.cell(row=row_num, column=5, value=stock_cumulatif)
        
        # Auto adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        # Generate response
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = f"mouvements_stock_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response

    @extend_schema(summary="Diagnostic de cohérence des stocks", tags=["Inventory"])
    @action(detail=False, methods=['get'], url_path='stock-diagnostic')
    def stock_diagnostic(self, request):
        """Vérifie la cohérence des stocks et identifie les incohérences."""
        from django.db.models import Sum
        from decimal import Decimal
        
        # 1. Statistiques des mouvements
        total_movements = StockMovement.objects.count()
        active_movements = StockMovement.objects.filter(is_active=True).count()
        inactive_movements = StockMovement.objects.filter(is_active=False).count()
        
        stats = {
            'totalMovements': total_movements,
            'activeMovements': active_movements,
            'inactiveMovements': inactive_movements,
        }
        
        # 2. Vérifier chaque produit/magasin
        stocks = Stock.objects.select_related('product', 'store').all()
        issues = []
        
        for stock in stocks:
            # Calculer le stock théorique basé sur les mouvements ACTIFS uniquement
            movements = StockMovement.objects.filter(
                product=stock.product,
                store=stock.store,
                is_active=True
            )
            
            entrees = movements.filter(movement_type='in').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            sorties = movements.filter(movement_type='out').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Transferts sortants
            transferts_out = movements.filter(movement_type='transfer').aggregate(
                total=Sum('quantity')
            )['total'] or Decimal('0')
            
            # Transferts entrants
            transferts_in = StockMovement.objects.filter(
                product=stock.product,
                destination_store=stock.store,
                movement_type='transfer',
                is_active=True
            ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
            
            stock_theorique = entrees - sorties - transferts_out + transferts_in
            stock_actuel = stock.quantity
            
            difference = stock_actuel - stock_theorique
            
            if abs(difference) > Decimal('0.01'):  # Tolérance pour les arrondis
                issues.append({
                    'product': stock.product.name,
                    'product_id': stock.product.id,
                    'store': stock.store.name,
                    'store_id': stock.store.id,
                    'stock_actuel': float(stock_actuel),
                    'stock_theorique': float(stock_theorique),
                    'difference': float(difference),
                    'entrees': float(entrees),
                    'sorties': float(sorties),
                    'transferts_out': float(transferts_out),
                    'transferts_in': float(transferts_in),
                })
        
        # 3. Bons d'entrée supprimés
        deleted_receipts = []
        inactive_receipts = StockMovement.objects.filter(
            is_active=False,
            receipt_number__isnull=False
        ).values('receipt_number').distinct()
        
        for receipt in inactive_receipts:
            receipt_num = receipt['receipt_number']
            movements = StockMovement.objects.filter(
                receipt_number=receipt_num,
                is_active=False
            )
            total_qty = movements.aggregate(total=Sum('quantity'))['total'] or 0
            deleted_receipts.append({
                'receipt_number': receipt_num,
                'count': movements.count(),
                'total_qty': float(total_qty)
            })
        
        return Response({
            'stats': stats,
            'issues': issues,
            'deleted_receipts': deleted_receipts,
        })
    
    @extend_schema(summary="Corriger les incohérences de stocks", tags=["Inventory"])
    @action(detail=False, methods=['post'], url_path='stock-diagnostic/fix')
    def fix_stock_diagnostic(self, request):
        """Corrige les incohérences de stocks détectées."""
        from django.db.models import Sum
        from decimal import Decimal
        from django.db import transaction
        
        corrected = 0
        errors = []
        
        # Récupérer toutes les incohérences
        stocks = Stock.objects.select_related('product', 'store').all()
        
        with transaction.atomic():
            for stock in stocks:
                # Calculer le stock théorique
                movements = StockMovement.objects.filter(
                    product=stock.product,
                    store=stock.store,
                    is_active=True
                )
                
                entrees = movements.filter(movement_type='in').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                sorties = movements.filter(movement_type='out').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                transferts_out = movements.filter(movement_type='transfer').aggregate(
                    total=Sum('quantity')
                )['total'] or Decimal('0')
                
                transferts_in = StockMovement.objects.filter(
                    product=stock.product,
                    destination_store=stock.store,
                    movement_type='transfer',
                    is_active=True
                ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
                
                stock_theorique = entrees - sorties - transferts_out + transferts_in
                stock_actuel = stock.quantity
                
                difference = stock_actuel - stock_theorique
                
                if abs(difference) > Decimal('0.01'):
                    try:
                        stock.quantity = stock_theorique
                        stock.save()
                        corrected += 1
                    except Exception as e:
                        errors.append({
                            'product': stock.product.name,
                            'store': stock.store.name,
                            'error': str(e)
                        })
        
        return Response({
            'corrected': corrected,
            'errors': errors,
            'message': f'{corrected} stock(s) corrigé(s) avec succès'
        })


@extend_schema_view(
    list=extend_schema(summary="Liste des transferts", tags=["Inventory"]),
    retrieve=extend_schema(summary="Détail d'un transfert", tags=["Inventory"]),
    create=extend_schema(summary="Créer un transfert", tags=["Inventory"]),
)
class StockTransferViewSet(viewsets.ModelViewSet):
    """ViewSet for StockTransfer model."""
    
    queryset = StockTransfer.objects.select_related(
        'source_store', 'destination_store'
    ).prefetch_related('lines__product')
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'inventory'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['source_store', 'destination_store', 'status']
    search_fields = ['transfer_number', 'source_store__name', 'destination_store__name', 'lines__product__name']
    ordering_fields = ['transfer_date', 'created_at']
    ordering = ['created_at']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                from django.db.models import Q
                return queryset.filter(
                    Q(source_store__in=user.assigned_stores.all()) |
                    Q(destination_store__in=user.assigned_stores.all())
                )
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        return queryset.filter(created_by=user)
    
    def get_serializer_class(self):
        if self.action == 'list':
            return StockTransferListSerializer
        elif self.action == 'create':
            return StockTransferCreateSerializer
        return StockTransferDetailSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def _can_validate_transfer(self, transfer):
        """Vérifier si l'utilisateur peut valider le transfert (envoi)."""
        user = self.request.user
        
        # Admin peut tout faire
        if user.is_superuser or user.role.name == 'super_admin':
            return True
        
        # Doit avoir la permission de gérer l'inventaire
        if not user.role.can_manage_inventory:
            return False
        
        # Doit être assigné au store source
        if user.role.access_scope == 'assigned':
            return transfer.source_store in user.assigned_stores.all()
        
        return True
    
    def _can_receive_transfer(self, transfer):
        """Vérifier si l'utilisateur peut recevoir le transfert."""
        user = self.request.user
        
        # Admin peut tout faire
        if user.is_superuser or user.role.name == 'super_admin':
            return True
        
        # Doit avoir la permission de voir/gérer l'inventaire
        if not (user.role.can_manage_inventory or user.role.can_view_inventory):
            return False
        
        # Doit être assigné au store destination
        if user.role.access_scope == 'assigned':
            return transfer.destination_store in user.assigned_stores.all()
        
        return True
    
    @extend_schema(summary="Valider le transfert", tags=["Inventory"])
    @action(detail=True, methods=['post'])
    def validate_transfer(self, request, pk=None):
        """Validate and send transfer."""
        transfer = self.get_object()
        
        # Vérifier les permissions
        if not self._can_validate_transfer(transfer):
            return Response(
                {'error': 'Vous n\'avez pas la permission de valider ce transfert. Vous devez être assigné au magasin source.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if transfer.status != 'draft':
            return Response(
                {'error': 'Seuls les transferts en brouillon peuvent être validés.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update quantities sent
        for line in transfer.lines.all():
            line.quantity_sent = line.quantity_requested
            line.save()
            
            # Decrease stock in source
            stock = Stock.objects.get(
                product=line.product,
                store=transfer.source_store
            )
            stock.quantity -= line.quantity_sent
            stock.save()
            
            # Créer un mouvement de stock de type "transfer" pour tracer la sortie
            StockMovement.objects.create(
                product=line.product,
                store=transfer.source_store,
                destination_store=transfer.destination_store,
                movement_type='transfer',
                quantity=line.quantity_sent,
                reference=transfer.transfer_number,
                notes=f'Transfert vers {transfer.destination_store.name}',
                date=transfer.transfer_date,
                created_by=request.user
            )
        
        transfer.status = 'in_transit'
        transfer.validated_by = request.user
        transfer.save()
        
        # Notifier les utilisateurs du magasin destination qu'un transfert est en route
        from core.notifications import create_notification
        from apps.accounts.models import User
        
        users_to_notify = User.objects.filter(
            is_active=True,
            role__isnull=False,
            assigned_stores=transfer.destination_store
        ).distinct()
        
        total_items = sum(line.quantity_sent for line in transfer.lines.all())
        
        for user in users_to_notify:
            create_notification(
                user=user,
                notification_type='transfer_in_transit',
                title='🚚 Transfert en route',
                message=f'Transfert {transfer.transfer_number} en transit : {int(total_items)} article(s) en provenance de "{transfer.source_store.name}" arrivent bientôt.',
                priority='medium',
                data={
                    'transfer_id': transfer.id,
                    'transfer_number': transfer.transfer_number,
                    'from_warehouse': transfer.source_store.name,
                    'total_items': int(total_items)
                },
                action_url=f'/inventory/transfers/{transfer.id}'
            )
        
        serializer = self.get_serializer(transfer)
        return Response(serializer.data)
    
    @extend_schema(summary="Recevoir le transfert", tags=["Inventory"])
    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Receive transfer."""
        transfer = self.get_object()
        
        # Vérifier les permissions
        if not self._can_receive_transfer(transfer):
            return Response(
                {'error': 'Vous n\'avez pas la permission de recevoir ce transfert. Vous devez être assigné au magasin destination.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if transfer.status != 'in_transit':
            return Response(
                {'error': 'Ce transfert ne peut pas être reçu.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Utiliser une transaction atomique pour éviter les problèmes
        with transaction.atomic():
            # Update quantities received
            for line in transfer.lines.all():
                line.quantity_received = line.quantity_sent
                line.save()
                
                # Increase stock in destination
                # Utiliser select_for_update pour éviter les conditions de course
                try:
                    stock = Stock.objects.select_for_update().get(
                        product=line.product,
                        store=transfer.destination_store
                    )
                    stock.quantity += line.quantity_received
                    stock.save()
                except Stock.DoesNotExist:
                    # Si le stock n'existe pas, créer directement avec la bonne quantité
                    # pour éviter le double save (0 puis +quantity) qui déclenche l'alerte
                    Stock.objects.create(
                        product=line.product,
                        store=transfer.destination_store,
                        quantity=line.quantity_received,
                        reserved_quantity=0
                    )
                
                # Créer un mouvement de stock de type "in" pour tracer l'entrée au magasin destination
                StockMovement.objects.create(
                    product=line.product,
                    store=transfer.destination_store,
                    movement_type='in',
                    quantity=line.quantity_received,
                    reference=transfer.transfer_number,
                    notes=f'Transfert reçu depuis {transfer.source_store.name}',
                    date=transfer.actual_arrival or timezone.now().date(),
                    created_by=request.user
                )
            
            transfer.status = 'received'
            transfer.received_by = request.user
            transfer.actual_arrival = timezone.now().date()
            transfer.save()
        
        # Notifier les utilisateurs du magasin destination
        from core.notifications import notify_transfer_received
        from apps.accounts.models import User
        
        # Obtenir les utilisateurs à notifier pour le magasin destination
        users_to_notify = User.objects.filter(
            is_active=True,
            role__isnull=False,
            assigned_stores=transfer.destination_store
        ).distinct()
        
        # Calculer le nombre total d'articles transférés
        total_items = sum(line.quantity_received for line in transfer.lines.all())
        
        # Envoyer la notification à chaque utilisateur
        for user in users_to_notify:
            notify_transfer_received(
                user=user,
                transfer_id=transfer.id,
                transfer_number=transfer.transfer_number,
                from_warehouse=transfer.source_store.name,
                to_warehouse=transfer.destination_store.name,
                total_items=int(total_items)
            )
        
        serializer = self.get_serializer(transfer)
        return Response(serializer.data)
    
    @extend_schema(summary="Annuler le transfert", tags=["Inventory"])
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel transfer."""
        transfer = self.get_object()
        
        if transfer.status not in ['draft', 'in_transit']:
            return Response(
                {'error': 'Ce transfert ne peut pas être annulé.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Si le transfert était en transit, remettre le stock au magasin source
        if transfer.status == 'in_transit':
            for line in transfer.lines.all():
                stock = Stock.objects.get(
                    product=line.product,
                    store=transfer.source_store
                )
                stock.quantity += line.quantity_sent
                stock.save()
        
        transfer.status = 'cancelled'
        transfer.save()
        
        serializer = self.get_serializer(transfer)
        return Response(serializer.data)
    
    @extend_schema(summary="Exporter les transferts en Excel", tags=["Inventory"])
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export transfers to Excel."""
        # Récupérer les filtres
        queryset = self.filter_queryset(self.get_queryset())
        
        # Appliquer les filtres de date si présents
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(transfer_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transfer_date__lte=end_date)
        
        transfers = queryset.select_related(
            'source_store', 'destination_store', 'created_by', 
            'validated_by', 'received_by'
        ).prefetch_related('lines__product')
        
        # Créer le workbook
        wb, ws = ExcelExporter.create_workbook("Transferts de Stock")
        
        # En-têtes
        columns = [
            'N° Transfert', 'Date', 'Magasin Source', 'Magasin Destination',
            'Nombre d\'articles', 'Statut', 'Date d\'arrivée prévue', 
            'Date d\'arrivée réelle', 'Créé par', 'Validé par', 'Reçu par'
        ]
        ExcelExporter.style_header(ws, columns)
        
        # Données
        status_labels = {
            'draft': 'Brouillon',
            'pending': 'En attente',
            'in_transit': 'En transit',
            'received': 'Reçu',
            'cancelled': 'Annulé'
        }
        
        for row_num, transfer in enumerate(transfers, 2):
            ws.cell(row=row_num, column=1, value=transfer.transfer_number)
            ws.cell(row=row_num, column=2, value=transfer.transfer_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=3, value=transfer.source_store.name)
            ws.cell(row=row_num, column=4, value=transfer.destination_store.name)
            ws.cell(row=row_num, column=5, value=transfer.lines.count())
            ws.cell(row=row_num, column=6, value=status_labels.get(transfer.status, transfer.status))
            ws.cell(row=row_num, column=7, value=transfer.expected_arrival.strftime('%d/%m/%Y') if transfer.expected_arrival else '')
            ws.cell(row=row_num, column=8, value=transfer.actual_arrival.strftime('%d/%m/%Y') if transfer.actual_arrival else '')
            ws.cell(row=row_num, column=9, value=transfer.created_by.get_full_name() if transfer.created_by else '')
            ws.cell(row=row_num, column=10, value=transfer.validated_by.get_full_name() if transfer.validated_by else '')
            ws.cell(row=row_num, column=11, value=transfer.received_by.get_full_name() if transfer.received_by else '')
        
        ExcelExporter.auto_adjust_columns(ws)
        
        filename = f"transferts_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExcelExporter.generate_response(wb, filename)


@extend_schema_view(
    list=extend_schema(summary="Liste des inventaires", tags=["Inventory"]),
    retrieve=extend_schema(summary="Détail d'un inventaire", tags=["Inventory"]),
    create=extend_schema(summary="Créer un inventaire", tags=["Inventory"]),
)
class InventoryViewSet(viewsets.ModelViewSet):
    """ViewSet for Inventory model."""
    
    queryset = Inventory.objects.select_related('store').prefetch_related('lines__product')
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'inventory'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['store', 'status']
    search_fields = ['inventory_number']
    ordering_fields = ['inventory_date', 'created_at']
    ordering = ['created_at']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                return queryset.filter(store__in=user.assigned_stores.all())
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        return queryset.filter(created_by=user)
    
    def get_serializer_class(self):
        if self.action == 'list':
            return InventoryListSerializer
        elif self.action == 'create':
            return InventoryCreateSerializer
        return InventoryDetailSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @extend_schema(summary="Valider l'inventaire", tags=["Inventory"])
    @action(detail=True, methods=['post'])
    def validate_inventory(self, request, pk=None):
        """Validate inventory and adjust stock."""
        inventory = self.get_object()
        
        if inventory.status != 'completed':
            return Response(
                {'error': 'Seuls les inventaires terminés peuvent être validés.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Adjust stock based on differences
        for line in inventory.lines.all():
            if line.difference != 0:
                stock = Stock.objects.get(
                    product=line.product,
                    store=inventory.store
                )
                stock.quantity = line.counted_quantity
                stock.save()
                
                # Create adjustment movement
                StockMovement.objects.create(
                    product=line.product,
                    store=inventory.store,
                    movement_type='adjustment',
                    quantity=abs(line.difference),
                    reference=inventory.inventory_number,
                    notes=f"Ajustement inventaire {inventory.inventory_number}",
                    created_by=request.user
                )
        
        inventory.status = 'validated'
        inventory.validated_by = request.user
        inventory.validation_date = timezone.now()
        inventory.save()
        
        serializer = self.get_serializer(inventory)
        return Response(serializer.data)
    
    @extend_schema(summary="Exporter l'état des stocks en Excel", tags=["Inventory"])
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export stock status to Excel."""
        stocks = Stock.objects.select_related('product', 'store').all()
        
        wb, ws = ExcelExporter.create_workbook("État des Stocks")
        
        columns = [
            'Produit', 'Référence', 'Magasin', 'Quantité',
            'Quantité Réservée', 'Quantité Disponible', 'Stock Min', 'Statut'
        ]
        ExcelExporter.style_header(ws, columns)
        
        for row_num, stock in enumerate(stocks, 2):
            ws.cell(row=row_num, column=1, value=stock.product.name)
            ws.cell(row=row_num, column=2, value=stock.product.reference)
            ws.cell(row=row_num, column=3, value=stock.store.name)
            ws.cell(row=row_num, column=4, value=float(stock.quantity))
            ws.cell(row=row_num, column=5, value=float(stock.reserved_quantity))
            ws.cell(row=row_num, column=6, value=float(stock.available_quantity))
            ws.cell(row=row_num, column=7, value=stock.product.minimum_stock)
            
            if stock.quantity < stock.product.minimum_stock:
                status = 'ALERTE'
            elif stock.quantity == 0:
                status = 'RUPTURE'
            else:
                status = 'OK'
            ws.cell(row=row_num, column=8, value=status)
        
        ExcelExporter.auto_adjust_columns(ws)
        
        filename = f"stocks_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExcelExporter.generate_response(wb, filename)
