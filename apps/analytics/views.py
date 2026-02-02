from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django.db.models import Sum, Count, Avg, F, Q, Case, When, DecimalField
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, datetime
from io import BytesIO

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.sales.models import Sale
from apps.products.models import Product
from apps.inventory.models import Stock
from apps.accounts.models import User
from apps.cashbox.models import CashMovement
from apps.loans.models import Loan
from apps.expenses.models import Expense


class DashboardViewSet(viewsets.ViewSet):
    """
    ViewSet for analytics and dashboard data with user-based filtering.
    - Super admin/Manager (access_scope='all'): voit toutes les statistiques
    - Caissier (access_scope='own'): voit uniquement ses propres statistiques
    """
    permission_classes = [IsAuthenticated]
    
    def _get_sales_queryset(self, user):
        """Helper to get filtered sales queryset based on user role."""
        queryset = Sale.objects.all()
        
        # Super admin voit tout
        if user.is_superuser:
            return queryset
        
        # Vérifier le scope d'accès du rôle
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                assigned_stores = user.assigned_stores.all()
                if assigned_stores.exists():
                    return queryset.filter(store__in=assigned_stores)
                else:
                    return queryset.none()
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        # Par défaut, filtrer par créateur
        return queryset.filter(created_by=user)
    
    def _get_expenses_queryset(self, user):
        """Helper to get filtered expenses queryset based on user role."""
        queryset = Expense.objects.all()
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                assigned_stores = user.assigned_stores.all()
                if assigned_stores.exists():
                    return queryset.filter(store__in=assigned_stores)
                else:
                    return queryset.none()
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        return queryset.filter(created_by=user)
    
    def _get_assigned_stores(self, user):
        """Helper to get user's assigned stores."""
        if user.is_superuser:
            return None  # Signifie pas de filtrage, voir tous les stores
        
        if hasattr(user, 'assigned_stores'):
            assigned_stores = user.assigned_stores.all()
            if assigned_stores.exists():
                return assigned_stores
        
        return None
    
    @action(detail=False, methods=['get'])
    def overview(self, request):
        """Main dashboard overview with user-specific data."""
        user = request.user
        today = timezone.now().date()
        month_start = today.replace(day=1)
        year_start = today.replace(month=1, day=1)
        
        # Récupérer le filtre de store (optionnel)
        store_filter = request.query_params.get('store')
        
        # Get assigned stores if user has any
        assigned_stores = self._get_assigned_stores(user)
        
        # Si un store est spécifié, filtrer uniquement sur ce store
        if store_filter:
            from apps.inventory.models import Store
            try:
                selected_store = Store.objects.get(id=store_filter, is_active=True)
                assigned_stores = [selected_store]
            except Store.DoesNotExist:
                assigned_stores = []
        
        # Get filtered querysets with optimizations
        sales_qs = self._get_sales_queryset(user).only(
            'id', 'sale_date', 'status', 'total_amount', 'payment_status'
        )
        expenses_qs = self._get_expenses_queryset(user).only(
            'id', 'status'
        )
        
        # Appliquer le filtre de store si spécifié
        if store_filter and assigned_stores:
            sales_qs = sales_qs.filter(store__in=assigned_stores)
            expenses_qs = expenses_qs.filter(store__in=assigned_stores)
        
        # Sales statistics - TOUTES LES VENTES (pas seulement confirmées)
        # Car dans la facturation, on peut avoir des ventes qui génèrent du CA sans être confirmées
        sales_today = sales_qs.filter(sale_date=today).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        
        sales_month = sales_qs.filter(
            sale_date__gte=month_start
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        
        sales_year = sales_qs.filter(
            sale_date__gte=year_start
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        
        # Stock statistics (optimized queries) - filtré par store si spécifié
        stock_queryset = Stock.objects.select_related('product')
        if store_filter and assigned_stores:
            stock_queryset = stock_queryset.filter(store__in=assigned_stores)
        
        stock_stats = {
            'total_products': Product.objects.filter(is_active=True).count(),
            'low_stock': stock_queryset.filter(
                quantity__lt=F('product__minimum_stock')
            ).only('id', 'product_id').distinct().count(),
            'out_of_stock': stock_queryset.filter(quantity=0).count(),
            'total_stock_value': stock_queryset.aggregate(
                total=Sum(F('quantity') * F('product__cost_price'))
            )['total'] or 0,
        }
        
        # Customer statistics - filtrées par les clients qui ont des ventes dans les stores assignés
        from apps.customers.models import Customer
        
        # Si l'utilisateur a des stores assignés, compter les clients de ces stores uniquement
        if assigned_stores:
            customer_queryset = Customer.objects.filter(
                sales__store__in=assigned_stores,
                sales__customer__isnull=False,
                is_active=True
            ).distinct()
            customer_stats = {
                'total': customer_queryset.count(),
                'new_this_month': Customer.objects.filter(
                    sales__store__in=assigned_stores,
                    sales__customer__isnull=False,
                    created_at__gte=month_start
                ).distinct().count(),
            }
        else:
            # Admin voit tous les clients
            customer_stats = {
                'total': Customer.objects.filter(is_active=True).count(),
                'new_this_month': Customer.objects.filter(
                    created_at__gte=month_start
                ).count(),
            }
        
        # Cash statistics - Utiliser la même fonction que l'API /cashbox/caisse/solde/
        from apps.cashbox.utils import get_cashbox_real_balance
        
        # Si un store est filtré, calculer uniquement pour ce store
        if store_filter and assigned_stores:
            cash_balance = sum(
                get_cashbox_real_balance(store_id=s.id) 
                for s in assigned_stores
            )
        elif user.is_superuser or (hasattr(user, 'role') and user.role and user.role.access_scope == 'all'):
            # Admin voit tout - solde total de tous les stores (ou store filtré)
            from apps.inventory.models import Store
            stores_to_calc = assigned_stores if assigned_stores else Store.objects.filter(is_active=True)
            cash_balance = sum(
                get_cashbox_real_balance(store_id=s.id) 
                for s in stores_to_calc
            )
        else:
            # Utilisateur normal voit uniquement ses stores assignés
            if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                cash_balance = sum(
                    get_cashbox_real_balance(store_id=s.id) 
                    for s in user.assigned_stores.all()
                )
            else:
                cash_balance = 0
        
        # Pending items (filtered by user)
        pending = {
            'sales': sales_qs.filter(status='draft').count(),
            'payments': sales_qs.filter(
                status='confirmed',
                payment_status__in=['unpaid', 'partial']
            ).count(),
            'expenses': expenses_qs.filter(status='pending').count(),
        }
        
        data = {
            'sales': {
                'today': {
                    'amount': float(sales_today['total'] or 0),
                    'count': sales_today['count']
                },
                'this_month': {
                    'amount': float(sales_month['total'] or 0),
                    'count': sales_month['count']
                },
                'this_year': {
                    'amount': float(sales_year['total'] or 0),
                    'count': sales_year['count']
                }
            },
            'stock': stock_stats,
            'customers': customer_stats,
            'cash': {
                'balance': float(cash_balance)
            },
            'pending': pending
        }
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def sales_chart(self, request):
        """Sales chart data over time (filtered by user)."""
        user = request.user
        period = request.query_params.get('period', 'month')  # week, month, year
        store_filter = request.query_params.get('store')  # Filtre par store
        
        today = timezone.now().date()
        
        if period == 'week':
            # Derniers 7 jours (une semaine)
            start_date = today - timedelta(days=7)
            trunc_func = TruncDate
        elif period == 'month':
            # Derniers 30 jours (un mois) - affichage jour par jour
            start_date = today - timedelta(days=30)
            trunc_func = TruncDate
        elif period == 'year':
            # Derniers 12 mois - affichage mois par mois
            start_date = today - timedelta(days=365)
            trunc_func = TruncMonth
        else:
            # Par défaut: dernier mois
            start_date = today - timedelta(days=30)
            trunc_func = TruncDate
        
        # Get filtered sales queryset
        sales_qs = self._get_sales_queryset(user)
        
        # Appliquer le filtre de store si spécifié
        if store_filter:
            sales_qs = sales_qs.filter(store_id=store_filter)
        
        sales_data = sales_qs.filter(
            sale_date__gte=start_date,
            status='confirmed'
        ).annotate(
            period=trunc_func('sale_date')
        ).values('period').annotate(
            total_amount=Sum('total_amount'),
            total_sales=Count('id')
        ).order_by('period')
        
        return Response(list(sales_data))
    
    @action(detail=False, methods=['get'])
    def top_products(self, request):
        """Top selling products (filtered by user) with real stock information."""
        user = request.user
        limit = int(request.query_params.get('limit', 10))
        
        from apps.sales.models import SaleLine
        from apps.inventory.models import Stock
        
        # Get filtered sales queryset
        sales_qs = self._get_sales_queryset(user)
        
        # Get top selling products
        top_products = SaleLine.objects.filter(
            sale__in=sales_qs,
            sale__status='confirmed',
            line_type='product'
        ).values(
            'product__id',
            'product__name',
            'product__reference'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum(F('quantity') * F('unit_price')),
            sales_count=Count('sale', distinct=True)
        ).order_by('-total_quantity')[:limit]  # Trier par quantité décroissante
        
        # Enrich with real stock data
        result = []
        for product_data in top_products:
            product_id = product_data['product__id']
            
            # Get stock for this product based on user's access
            if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.access_scope == 'all'):
                # Admin sees total stock across all stores
                total_stock = Stock.objects.filter(
                    product_id=product_id
                ).aggregate(total=Sum('quantity'))['total'] or 0
            else:
                # Users with assigned stores see stock in their stores
                assigned_stores = user.assigned_stores.all() if hasattr(user, 'assigned_stores') else []
                if assigned_stores.exists():
                    total_stock = Stock.objects.filter(
                        product_id=product_id,
                        store__in=assigned_stores
                    ).aggregate(total=Sum('quantity'))['total'] or 0
                else:
                    total_stock = 0
            
            result.append({
                'product__id': product_data['product__id'],
                'product__name': product_data['product__name'],
                'product__reference': product_data['product__reference'],
                'total_quantity': product_data['total_quantity'],
                'total_amount': float(product_data['total_amount']),
                'sales_count': product_data['sales_count'],
                'current_stock': total_stock  # Stock réel
            })
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def top_customers(self, request):
        """Top customers by revenue (filtered by user)."""
        user = request.user
        limit = int(request.query_params.get('limit', 10))
        
        # Get filtered sales queryset
        sales_qs = self._get_sales_queryset(user)
        
        top_customers = sales_qs.filter(
            status='confirmed',
            customer__isnull=False  # Exclure les ventes sans client
        ).values(
            'customer__id',
            'customer__name',
            'customer__customer_code'
        ).annotate(
            total_amount=Sum('total_amount'),
            sales_count=Count('id')
        ).order_by('total_amount')[:limit]
        
        # Calculate average order value manually after aggregation
        result = []
        for customer in top_customers:
            customer_data = dict(customer)
            customer_data['avg_order'] = float(customer['total_amount'] / customer['sales_count']) if customer['sales_count'] > 0 else 0
            customer_data['total_amount'] = float(customer['total_amount'])
            result.append(customer_data)
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def revenue_by_category(self, request):
        """Revenue breakdown by product category."""
        from apps.sales.models import SaleLine
        
        category_revenue = SaleLine.objects.filter(
            sale__status='confirmed',
            line_type='product'
        ).values(
            'product__category__id',
            'product__category__name'
        ).annotate(
            total_revenue=Sum(F('quantity') * F('unit_price')),
            total_quantity=Sum('quantity')
        ).order_by('total_revenue')
        
        return Response(list(category_revenue))
    
    @action(detail=False, methods=['get'])
    def cash_flow(self, request):
        """Cash flow analysis."""
        days = int(request.query_params.get('days', 30))
        store_filter = request.query_params.get('store')  # Filtre par store
        start_date = timezone.now().date() - timedelta(days=days)
        
        cash_movements = CashMovement.objects.filter(
            created_at__date__gte=start_date
        )
        
        # Appliquer le filtre de store si spécifié
        # CashMovement -> cashbox_session -> cashbox -> store
        if store_filter:
            cash_movements = cash_movements.filter(
                cashbox_session__cashbox__store_id=store_filter
            )
        
        cash_movements = cash_movements.annotate(
            date=TruncDate('created_at')
        ).values('date', 'movement_type').annotate(
            total=Sum('amount')
        ).order_by('date')
        
        # Structure data by date
        result = {}
        for movement in cash_movements:
            date_str = str(movement['date'])
            if date_str not in result:
                result[date_str] = {'date': date_str, 'in': 0, 'out': 0}
            
            if movement['movement_type'] == 'in':
                result[date_str]['in'] = float(movement['total'])
            else:
                result[date_str]['out'] = float(movement['total'])
        
        return Response(list(result.values()))
    
    @action(detail=False, methods=['get'])
    def inventory_value(self, request):
        """Inventory value by store."""
        inventory_by_store = Stock.objects.values(
            'store__id',
            'store__name'
        ).annotate(
            total_value=Sum(F('quantity') * F('product__cost_price')),
            total_products=Count('product', distinct=True),
            total_quantity=Sum('quantity')
        ).order_by('total_value')
        
        return Response(list(inventory_by_store))
    
    @action(detail=False, methods=['get'])
    def financial_summary(self, request):
        """Financial summary including loans, expenses, etc. (filtered by user)."""
        user = request.user
        month_start = timezone.now().date().replace(day=1)
        store_filter = request.query_params.get('store')  # Filtre par store
        
        # Get filtered querysets
        expenses_qs = self._get_expenses_queryset(user)
        sales_qs = self._get_sales_queryset(user)
        
        # Appliquer le filtre de store si spécifié
        if store_filter:
            expenses_qs = expenses_qs.filter(store_id=store_filter)
            sales_qs = sales_qs.filter(store_id=store_filter)
        
        # Loans - toujours global car pas de système de propriété
        loans_summary = {
            'total_principal': Loan.objects.aggregate(
                total=Sum('principal_amount')
            )['total'] or 0,
            'total_debt': Loan.objects.filter(
                status='active'
            ).aggregate(
                total=Sum(F('total_amount') - F('paid_amount'))
            )['total'] or 0,
            'monthly_payment': Loan.objects.filter(
                status='active'
            ).aggregate(
                total=Sum(F('total_amount') / F('duration_months'))
            )['total'] or 0,
        }
        
        # Expenses - FILTRÉ par utilisateur
        expenses_summary = {
            'this_month': expenses_qs.filter(
                expense_date__gte=month_start,
                status='paid'
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'pending': expenses_qs.filter(
                status__in=['pending', 'approved']
            ).aggregate(total=Sum('amount'))['total'] or 0,
        }
        
        # Revenue - FILTRÉ par utilisateur
        revenue_summary = {
            'this_month': sales_qs.filter(
                sale_date__gte=month_start,
                status='confirmed'
            ).aggregate(total=Sum('total_amount'))['total'] or 0,
        }
        
        # Calculate profit
        profit = float(revenue_summary['this_month']) - float(expenses_summary['this_month'])
        
        return Response({
            'loans': loans_summary,
            'expenses': expenses_summary,
            'revenue': revenue_summary,
            'profit': profit
        })
    
    @action(detail=False, methods=['get'], url_path='reporting-stats')
    def reporting_stats(self, request):
        """Get statistics for reporting page - Encaissements réels."""
        from apps.invoicing.models import Invoice, InvoicePayment
        
        # Total factures émises
        total_invoices = Invoice.objects.count()
        
        # Total expenses (paid and approved)
        total_expenses = Expense.objects.filter(
            Q(status='paid') | Q(status='approved')
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Total ENCAISSEMENTS (argent réellement reçu)
        # 1. Paiements de factures
        total_invoice_payments = InvoicePayment.objects.filter(
            status='success'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # 2. Montants payés des ventes directes
        total_sales_paid = Sale.objects.filter(
            status__in=['confirmed', 'completed']
        ).aggregate(total=Sum('paid_amount'))['total'] or 0
        
        total_sales = float(total_invoice_payments or 0) + float(total_sales_paid or 0)
        
        return Response({
            'total_invoices': total_invoices,
            'total_expenses': float(total_expenses),
            'total_sales': total_sales
        })
    
    @action(detail=False, methods=['get'], url_path='generate-report-data')
    def generate_report_data(self, request):
        """Generate report data with expenses and sales list."""
        from django.db.models import Sum
        
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        if not start_date_str or not end_date_str:
            return Response({'error': 'Les dates sont requises'}, status=400)
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        # Récupérer les charges
        expenses = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date,
            status__in=['paid', 'approved']
        ).select_related('category').order_by('expense_date')
        
        # Regrouper les charges par catégorie
        expenses_by_category = {}
        for exp in expenses:
            category_name = exp.category.name if exp.category else 'Divers'
            
            if category_name not in expenses_by_category:
                expenses_by_category[category_name] = {'total': 0, 'count': 0}
            
            expenses_by_category[category_name]['total'] += float(exp.amount)
            expenses_by_category[category_name]['count'] += 1
        
        # Ajouter les emprunts comme charges
        from apps.loans.models import Loan
        loans = Loan.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date,
            status__in=['active', 'paid']
        ).order_by('start_date')
        
        if loans.exists():
            if 'Emprunts' not in expenses_by_category:
                expenses_by_category['Emprunts'] = {'total': 0, 'count': 0}
            
            for loan in loans:
                expenses_by_category['Emprunts']['total'] += float(loan.total_amount)
                expenses_by_category['Emprunts']['count'] += 1
        
        # Convertir en liste
        expenses_data = [
            {
                'category_name': category,
                'amount': data['total'],
                'count': data['count']
            }
            for category, data in expenses_by_category.items()
        ]
        
        # Récupérer les ENCAISSEMENTS RÉELS (pas le CA)
        # 1. Paiements de factures
        from apps.invoicing.models import InvoicePayment
        invoice_payments = InvoicePayment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            status='success'
        ).select_related('invoice', 'invoice__customer').order_by('payment_date')
        
        # 2. Ventes avec montant payé
        sales = Sale.objects.filter(
            sale_date__gte=start_date,
            sale_date__lte=end_date,
            status__in=['confirmed', 'completed'],
            paid_amount__gt=0
        ).select_related('customer').order_by('sale_date')
        
        # Regrouper les ventes par catégorie
        sales_by_category = {}
        
        # Ajouter les paiements de factures
        for payment in invoice_payments:
            # Parcourir TOUTES les lignes de la facture pour extraire les catégories
            if hasattr(payment.invoice, 'lines'):
                for line in payment.invoice.lines.all():
                    category_name = 'Ventes diverses'
                    line_amount = float(line.total) if hasattr(line, 'total') else 0
                    
                    if hasattr(line, 'product') and line.product:
                        if hasattr(line.product, 'category') and line.product.category:
                            category_name = line.product.category.name
                    elif hasattr(line, 'service') and line.service:
                        if hasattr(line.service, 'category') and line.service.category:
                            category_name = line.service.category.name
                    
                    if category_name not in sales_by_category:
                        sales_by_category[category_name] = 0
                    
                    # Calculer la proportion de ce produit/service dans la facture
                    invoice_total = float(payment.invoice.total_amount)
                    proportion = line_amount / invoice_total if invoice_total > 0 else 0
                    amount_to_add = float(payment.amount) * proportion
                    
                    sales_by_category[category_name] += amount_to_add
        
        # Ajouter les ventes directes (montant payé)
        for sale in sales:
            # Parcourir TOUTES les lignes de la vente
            if hasattr(sale, 'lines'):
                for line in sale.lines.all():
                    category_name = 'Ventes diverses'
                    line_total = float(line.quantity) * float(line.unit_price)
                    
                    if hasattr(line, 'product') and line.product:
                        if hasattr(line.product, 'category') and line.product.category:
                            category_name = line.product.category.name
                    elif hasattr(line, 'service') and line.service:
                        if hasattr(line.service, 'category') and line.service.category:
                            category_name = line.service.category.name
                    
                    if category_name not in sales_by_category:
                        sales_by_category[category_name] = 0
                    
                    # Calculer la proportion de ce produit/service dans la vente
                    sale_total = float(sale.total_amount)
                    proportion = line_total / sale_total if sale_total > 0 else 0
                    amount_to_add = float(sale.paid_amount) * proportion
                    
                    sales_by_category[category_name] += amount_to_add
        
        # Convertir en liste pour le frontend
        sales_data = [
            {
                'category_name': category,
                'amount': total
            }
            for category, total in sales_by_category.items()
        ]
        
        return Response({
            'expenses': expenses_data,
            'sales': sales_data
        })
    
    @action(detail=False, methods=['post'], url_path='export-report')
    def export_report(self, request):
        """Export report as PDF or Excel file with expenses and sales grouped by category."""
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')
        format_type = request.data.get('format', 'pdf')
        orientation = request.data.get('orientation', 'landscape')
        page_size = request.data.get('page_size', 'A4')
        group_by_category = request.data.get('group_by_category', True)
        
        if not start_date_str or not end_date_str:
            return Response({'error': 'Les dates sont requises'}, status=400)
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        # Récupérer les charges
        expenses = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date,
            status__in=['paid', 'approved']
        ).select_related('category')
        
        # Grouper les charges par catégorie
        expenses_by_category = {}
        total_expenses = 0
        for exp in expenses:
            category_name = exp.category.name if exp.category else 'Divers'
            amount = float(exp.amount)
            if category_name not in expenses_by_category:
                expenses_by_category[category_name] = {'amount': 0, 'count': 0}
            expenses_by_category[category_name]['amount'] += amount
            expenses_by_category[category_name]['count'] += 1
            total_expenses += amount
        
        # Ajouter les emprunts comme charges
        from apps.loans.models import Loan
        loans = Loan.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date,
            status__in=['active', 'paid']
        )
        
        if loans.exists():
            if 'Emprunts' not in expenses_by_category:
                expenses_by_category['Emprunts'] = {'amount': 0, 'count': 0}
            
            for loan in loans:
                loan_amount = float(loan.total_amount)
                expenses_by_category['Emprunts']['amount'] += loan_amount
                expenses_by_category['Emprunts']['count'] += 1
                total_expenses += loan_amount
        
        # Récupérer les ENCAISSEMENTS RÉELS (même logique que generate-report-data)
        from apps.invoicing.models import InvoicePayment
        
        # Grouper les ventes par catégorie de produit/service
        sales_by_category = {}
        total_sales = 0
        
        # 1. Paiements de factures
        invoice_payments = InvoicePayment.objects.filter(
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            status='success'
        ).select_related('invoice').prefetch_related('invoice__lines__product__category', 'invoice__lines__service__category')
        
        for payment in invoice_payments:
            total_sales += float(payment.amount)
            
            # Parcourir TOUTES les lignes de la facture pour extraire les catégories
            if hasattr(payment.invoice, 'lines'):
                for line in payment.invoice.lines.all():
                    category_name = 'Ventes diverses'
                    line_amount = float(line.total) if hasattr(line, 'total') else 0
                    
                    # Récupérer la catégorie du produit ou du service
                    if hasattr(line, 'product') and line.product:
                        if hasattr(line.product, 'category') and line.product.category:
                            category_name = line.product.category.name
                    elif hasattr(line, 'service') and line.service:
                        if hasattr(line.service, 'category') and line.service.category:
                            category_name = line.service.category.name
                    
                    if category_name not in sales_by_category:
                        sales_by_category[category_name] = {'amount': 0, 'count': 0}
                    
                    # Calculer la proportion de ce produit/service dans la facture
                    invoice_total = float(payment.invoice.total_amount)
                    proportion = line_amount / invoice_total if invoice_total > 0 else 0
                    amount_to_add = float(payment.amount) * proportion
                    
                    sales_by_category[category_name]['amount'] += amount_to_add
                    sales_by_category[category_name]['count'] += 1
        
        # 2. Ventes directes avec montant payé
        sales = Sale.objects.filter(
            sale_date__gte=start_date,
            sale_date__lte=end_date,
            status__in=['confirmed', 'completed'],
            paid_amount__gt=0
        ).prefetch_related('lines__product__category', 'lines__service__category')
        
        for sale in sales:
            total_sales += float(sale.paid_amount)
            
            # Parcourir TOUTES les lignes de la vente pour extraire les catégories
            if hasattr(sale, 'lines'):
                for line in sale.lines.all():
                    category_name = 'Ventes diverses'
                    line_total = float(line.quantity) * float(line.unit_price)
                    
                    # Récupérer la catégorie du produit ou du service
                    if hasattr(line, 'product') and line.product:
                        if hasattr(line.product, 'category') and line.product.category:
                            category_name = line.product.category.name
                    elif hasattr(line, 'service') and line.service:
                        if hasattr(line.service, 'category') and line.service.category:
                            category_name = line.service.category.name
                    
                    if category_name not in sales_by_category:
                        sales_by_category[category_name] = {'amount': 0, 'count': 0}
                    
                    # Calculer la proportion de ce produit/service dans la vente
                    sale_total = float(sale.total_amount)
                    proportion = line_total / sale_total if sale_total > 0 else 0
                    amount_to_add = float(sale.paid_amount) * proportion
                    
                    sales_by_category[category_name]['amount'] += amount_to_add
                    sales_by_category[category_name]['count'] += 1
        
        # Calculer le résultat
        net_profit = total_sales - total_expenses
        
        period = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        
        # Generate PDF
        if format_type == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            
            buffer = BytesIO()
            # Utiliser le format paysage
            page_format = landscape(A4) if orientation == 'landscape' else A4
            doc = SimpleDocTemplate(buffer, pagesize=page_format, rightMargin=40, leftMargin=40, topMargin=30, bottomMargin=30)
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=28,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=3,
                fontName='Helvetica-Bold',
                alignment=0
            )
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#6b7280'),
                spaceAfter=15,
                alignment=0
            )
            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading2'],
                fontSize=10,
                textColor=colors.HexColor('#374151'),
                spaceAfter=8,
                spaceBefore=10,
                fontName='Helvetica-Bold'
            )
            
            # En-tête avec période
            header_data = [
                [Paragraph("Compte de Résultat", title_style), Paragraph(f"PÉRIODE<br/>{period}<br/>Généré le {datetime.now().strftime('%d/%m/%Y')}", 
                    ParagraphStyle('HeaderRight', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#9ca3af'), alignment=2))]
            ]
            header_table = Table(header_data, colWidths=[400, 300])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 10))
            
            # Ligne de séparation
            line = Table([['']], colWidths=[page_format[0] - 80])
            line.setStyle(TableStyle([
                ('LINEABOVE', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ]))
            elements.append(line)
            elements.append(Spacer(1, 15))
            
            # Cartes de résumé
            result_label = "Bénéfice" if net_profit >= 0 else "Perte"
            summary_data = [
                [
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='9' color='#6b7280'>VENTES</font></para>", styles['Normal']),
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='9' color='#6b7280'>CHARGES</font></para>", styles['Normal']),
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='9' color='#1f2937'><b>RÉSULTAT</b></font></para>", styles['Normal'])
                ],
                [
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='24'><b>{total_sales:,.0f}</b></font></para>", styles['Normal']),
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='24'><b>{total_expenses:,.0f}</b></font></para>", styles['Normal']),
                    Paragraph(f"<para align='center' spaceAfter='0'><font size='24'><b>{abs(net_profit):,.0f}</b></font></para>", styles['Normal'])
                ],
                [
                    Paragraph(f"<para align='center'><font size='8' color='#9ca3af'>FCFA</font></para>", styles['Normal']),
                    Paragraph(f"<para align='center'><font size='8' color='#9ca3af'>FCFA</font></para>", styles['Normal']),
                    Paragraph(f"<para align='center'><font size='8' color='#6b7280'>{result_label}</font></para>", styles['Normal'])
                ]
            ]
            summary_table = Table(summary_data, colWidths=[230, 230, 230])
            summary_table.setStyle(TableStyle([
                # Première ligne - labels
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#f3f4f6')),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#f3f4f6')),
                ('BACKGROUND', (2, 0), (2, 0), colors.white),
                ('BOX', (0, 0), (0, 2), 1, colors.HexColor('#e5e7eb')),
                ('BOX', (1, 0), (1, 2), 1, colors.HexColor('#e5e7eb')),
                ('BOX', (2, 0), (2, 2), 2, colors.HexColor('#1f2937')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 1), (-1, 1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, 1), 5),
                ('TOPPADDING', (0, 2), (-1, 2), 5),
                ('BOTTOMPADDING', (0, 2), (-1, 2), 10),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 15))
            
            # Tableau des Ventes groupées par catégorie
            elements.append(Paragraph("PRODUITS D'EXPLOITATION", section_style))
            sales_data = [['Catégorie', 'Montant']]
            for category, data in sorted(sales_by_category.items()):
                sales_data.append([category, f"{data['amount']:,.0f}"])
            sales_data.append(['TOTAL PRODUITS', f"{total_sales:,.0f}"])
            
            sales_table = Table(sales_data, colWidths=[540, 150])
            sales_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -2), 6),
                ('TOPPADDING', (0, 0), (-1, -2), 6),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#d1d5db')),
                ('LINEBELOW', (0, 1), (-1, -2), 0.5, colors.HexColor('#f3f4f6')),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f2937')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1f2937')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('TOPPADDING', (0, -1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            ]))
            elements.append(sales_table)
            elements.append(Spacer(1, 15))
            
            # Tableau des Charges groupées par catégorie
            elements.append(Paragraph("CHARGES D'EXPLOITATION", section_style))
            expenses_data = [['Catégorie', 'Montant']]
            for category, data in sorted(expenses_by_category.items()):
                expenses_data.append([category, f"{data['amount']:,.0f}"])
            expenses_data.append(['TOTAL CHARGES', f"{total_expenses:,.0f}"])
            
            expenses_table = Table(expenses_data, colWidths=[540, 150])
            expenses_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6b7280')),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -2), 6),
                ('TOPPADDING', (0, 0), (-1, -2), 6),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#d1d5db')),
                ('LINEBELOW', (0, 1), (-1, -2), 0.5, colors.HexColor('#f3f4f6')),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f2937')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1f2937')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('TOPPADDING', (0, -1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            ]))
            elements.append(expenses_table)
            elements.append(Spacer(1, 15))
            
            # Statistiques finales
            stats_data = [[
                Paragraph(f"<para align='center'><font size='8' color='#6b7280'>Catégories Produits</font><br/><font size='16'>{len(sales_by_category)}</font></para>", styles['Normal']),
                Paragraph(f"<para align='center'><font size='8' color='#6b7280'>Catégories Charges</font><br/><font size='16'>{len(expenses_by_category)}</font></para>", styles['Normal']),
                Paragraph(f"<para align='center'><font size='8' color='#6b7280'>Marge Nette</font><br/><font size='16'>{((net_profit / total_sales) * 100) if total_sales > 0 else 0:.1f}%</font></para>", styles['Normal'])
            ]]
            stats_table = Table(stats_data, colWidths=[230, 230, 230])
            stats_table.setStyle(TableStyle([
                ('LINEABOVE', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(stats_table)
            
            # Pied de page
            footer_line = Table([['']], colWidths=[page_format[0] - 80])
            footer_line.setStyle(TableStyle([
                ('LINEABOVE', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ]))
            elements.append(Spacer(1, 10))
            elements.append(footer_line)
            elements.append(Spacer(1, 5))
            footer_text = Paragraph(f"<para align='center'><font size='7' color='#9ca3af'>Document généré le {datetime.now().strftime('%d %B %Y')}</font></para>", styles['Normal'])
            elements.append(footer_text)
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="rapport_{start_date}_{end_date}.pdf"'
            return response
        
        # Generate Excel
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporting"
            
            # Styles
            title_font = Font(name='Arial', size=16, bold=True)
            subtitle_font = Font(name='Arial', size=10, color='666666')
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
            total_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
            total_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            info_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
            info_font = Font(name='Arial', size=9, color='FFFFFF')
            
            row = 1
            
            # Titre
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "Reporting Mensuel"
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Période
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = f"Période {period}"
            cell.font = subtitle_font
            cell.alignment = Alignment(horizontal='center')
            row += 2
            
            # Charges
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "Charges d'exploitation"
            cell.font = header_font
            cell.fill = header_fill
            row += 1
            
            # Headers Charges
            headers_charges = ['Catégorie', 'Qté', 'Montant']
            for col_idx, header in enumerate(headers_charges, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data Charges
            for category, data in sorted(expenses_by_category.items()):
                ws.cell(row=row, column=1, value=category)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=data['amount']).number_format = '#,##0'
                row += 1
            
            # Total Charges
            ws.cell(row=row, column=2, value="Total Charges").font = total_font
            ws.cell(row=row, column=2).fill = total_fill
            ws.cell(row=row, column=3, value=total_expenses).number_format = '#,##0'
            ws.cell(row=row, column=3).font = total_font
            ws.cell(row=row, column=3).fill = total_fill
            row += 2
            
            # Ventes
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "Produits d'exploitation"
            cell.font = header_font
            cell.fill = header_fill
            row += 1
            
            # Headers Ventes
            headers_ventes = ['Catégorie', 'Qté', 'Montant']
            for col_idx, header in enumerate(headers_ventes, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data Ventes
            for category, data in sorted(sales_by_category.items()):
                ws.cell(row=row, column=1, value=category)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=data['amount']).number_format = '#,##0'
                row += 1
            
            # Total Ventes
            ws.cell(row=row, column=2, value="Total Produits").font = total_font
            ws.cell(row=row, column=2).fill = total_fill
            ws.cell(row=row, column=3, value=total_sales).number_format = '#,##0'
            ws.cell(row=row, column=3).font = total_font
            ws.cell(row=row, column=3).fill = total_fill
            row += 2
            
            # Résultat
            result_color = '10b981' if net_profit >= 0 else 'ef4444'
            result_fill = PatternFill(start_color=result_color, end_color=result_color, fill_type='solid')
            result_font_styled = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws[f'A{row}']
            cell.value = "Résultat de la période"
            cell.font = Font(name='Arial', size=12, bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Bénéfice")
            ws.cell(row=row, column=2, value=f"{net_profit:,.0f} FCFA").font = result_font_styled
            ws.cell(row=row, column=2).fill = result_fill
            row += 2
            
            # Informations complémentaires
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "Informations complémentaires"
            cell.font = info_font
            cell.fill = info_fill
            row += 1
            
            info_lines = [
                f"Nombre de catégories de charges: {len(expenses_by_category)}",
                f"Nombre de catégories de produits: {len(sales_by_category)}",
                f"Ce rapport est établi pour la période {period}. La période représente un bénéfice de {net_profit:,.0f} FCFA."
            ]
            for info in info_lines:
                ws.merge_cells(f'A{row}:C{row}')
                cell = ws[f'A{row}']
                cell.value = info
                cell.font = info_font
                cell.fill = info_fill
                row += 1
            
            # Ajuster les largeurs de colonnes
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="rapport_{start_date}_{end_date}.xlsx"'
            return response

