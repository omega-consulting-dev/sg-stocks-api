from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Sum, F, Value, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
import io

from core.utils.export_utils import ExcelExporter
from apps.suppliers.models import Supplier, SupplierPayment
from apps.suppliers.serializers import (
    SupplierListSerializer,
    SupplierDetailSerializer,
    SupplierCreateUpdateSerializer,
    SupplierPaymentSerializer
)
from apps.suppliers.filters import SupplierFilter


class SupplierViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Supplier management with role-based filtering.
    - Super admin / Manager (access_scope='all'): voit tous les fournisseurs
    - Magasinier (access_scope='assigned'): voit les fournisseurs de ses stores
    - Caissier (access_scope='own'): voit uniquement les fournisseurs qu'il a créés ou avec qui il a fait des transactions
    """
    queryset = Supplier.objects.filter(is_active=True)
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = SupplierFilter
    search_fields = ['supplier_code', 'name', 'contact_person', 'email', 'phone', 'mobile']
    ordering_fields = ['name', 'supplier_code', 'created_at', 'rating', 'city']
    ordering = ['name']
    
    def get_queryset(self):
        """
        Filtrage sécurisé des fournisseurs selon le rôle et access_scope.
        """
        queryset = super().get_queryset()
        user = self.request.user
        
        # Super admin voit tout
        if user.is_superuser:
            return queryset
        
        # Vérifier le scope d'accès du rôle
        if hasattr(user, 'role') and user.role:
            # Manager avec accès à tous les fournisseurs
            if user.role.access_scope == 'all':
                return queryset
            
            # Magasinier: fournisseurs des stores assignés
            elif user.role.access_scope == 'assigned':
                from apps.suppliers.models import PurchaseOrder
                assigned_stores = user.assigned_stores.all()
                
                if not assigned_stores.exists():
                    return queryset.filter(created_by=user)
                    
                # Fournisseurs avec commandes dans les stores assignés
                supplier_ids_from_po = PurchaseOrder.objects.filter(
                    store__in=assigned_stores
                ).values_list('supplier_id', flat=True).distinct()
                
                # Union des fournisseurs + ceux créés par l'utilisateur
                return queryset.filter(
                    Q(id__in=supplier_ids_from_po) |
                    Q(created_by=user)
                ).distinct()
            
            # Caissier: uniquement fournisseurs créés par lui ou avec qui il a fait des transactions
            elif user.role.access_scope == 'own':
                from apps.suppliers.models import PurchaseOrder
                
                # Fournisseurs des commandes créées par le caissier
                supplier_ids_from_po = PurchaseOrder.objects.filter(
                    created_by=user
                ).values_list('supplier_id', flat=True).distinct()
                
                # Union des fournisseurs + ceux créés directement par le caissier
                return queryset.filter(
                    Q(id__in=supplier_ids_from_po) |
                    Q(created_by=user)
                ).distinct()
        
        # Par défaut, filtrer par créateur (sécurité)
        return queryset.filter(created_by=user)
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return SupplierListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return SupplierCreateUpdateSerializer
        return SupplierDetailSerializer
    
    def perform_create(self, serializer):
        """Enregistrer le créateur du fournisseur."""
        # Vérifier la permission
        user = self.request.user
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if not user.role.can_manage_suppliers:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("Vous n'avez pas la permission de créer des fournisseurs.")
        
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        """Vérifier la permission avant la mise à jour."""
        user = self.request.user
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if not user.role.can_manage_suppliers:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("Vous n'avez pas la permission de modifier des fournisseurs.")
        
        serializer.save()
    
    def perform_destroy(self, instance):
        """Soft delete: désactiver au lieu de supprimer."""
        # Vérifier la permission
        user = self.request.user
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if not user.role.can_manage_suppliers:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("Vous n'avez pas la permission de supprimer des fournisseurs.")
        
        instance.is_active = False
        instance.save()
    
    @action(detail=False, methods=['get'])
    def debts(self, request):
        """
        Liste les fournisseurs avec solde dû pour le tenant courant.
        GET /api/v1/suppliers/debts/
        Retourne les fournisseurs liés à des PurchaseOrder avec un solde non réglé.
        Les données sont automatiquement filtrées par access_scope (own/assigned/all).
        """
        # Utiliser get_queryset() pour respecter le filtrage par access_scope
        suppliers = self.get_queryset()
        
        data = []
        for supplier in suppliers:
            balance = supplier.get_balance()
            if balance > 0:
                # Calculer total_ordered et total_paid pour les PO avec balance > 0
                statuses = ['confirmed', 'received']
                pos_with_debt = supplier.purchase_orders.filter(
                    status__in=statuses
                ).annotate(
                    balance_calc=F('total_amount') - F('paid_amount')
                ).filter(balance_calc__gt=0)
                
                total_ordered = sum(po.total_amount for po in pos_with_debt)
                total_paid = sum(po.paid_amount for po in pos_with_debt)
                
                data.append({
                    'id': supplier.id,
                    'supplier_code': supplier.supplier_code,
                    'name': supplier.name,
                    'email': supplier.email,
                    'phone': supplier.phone,
                    'total_ordered': float(total_ordered),
                    'total_paid': float(total_paid),
                    'balance': float(balance),
                })
        
        # Trier par balance décroissante
        data.sort(key=lambda x: x['balance'], reverse=True)
        return Response(data)

    @action(detail=True, methods=['get'], url_path='purchase-orders-with-debt')
    def purchase_orders_with_debt(self, request, pk=None):
        """
        Retourne la liste des purchase orders avec dette pour un fournisseur spécifique.
        GET /api/v1/suppliers/{id}/purchase-orders-with-debt/
        """
        supplier = self.get_object()
        
        # Récupérer les commandes confirmées ou reçues avec un solde dû
        statuses = ['confirmed', 'received']
        purchase_orders = supplier.purchase_orders.filter(
            status__in=statuses
        ).annotate(
            balance_calc=F('total_amount') - F('paid_amount')
        ).filter(balance_calc__gt=0).order_by('order_date')
        
        data = []
        for po in purchase_orders:
            # Utiliser actual_delivery si disponible, sinon expected_delivery
            delivery_date = po.actual_delivery if po.actual_delivery else po.expected_delivery
            
            data.append({
                'id': po.id,
                'order_number': po.order_number,
                'order_date': po.order_date.isoformat() if po.order_date else None,
                'delivery_date': delivery_date.isoformat() if delivery_date else None,
                'due_date': po.due_date.isoformat() if po.due_date else None,
                'total_amount': float(po.total_amount),
                'paid_amount': float(po.paid_amount),
                'balance_due': float(po.total_amount - po.paid_amount),
                'status': po.status
            })
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export suppliers to Excel with filters support."""
        # Use filtered queryset
        suppliers = self.filter_queryset(self.get_queryset())
        
        wb, ws = ExcelExporter.create_workbook("Fournisseurs")
        
        columns = [
            'Code Fournisseur', 'Nom', 'Contact Principal', 'Email', 
            'Téléphone', 'Mobile', 'Ville', 'Pays', 'Conditions Paiement',
            'Évaluation', 'Solde Dû', 'Actif', 'Date Création'
        ]
        ExcelExporter.style_header(ws, columns)
        
        for row_num, supplier in enumerate(suppliers, 2):
            ws.cell(row=row_num, column=1, value=supplier.supplier_code)
            ws.cell(row=row_num, column=2, value=supplier.name)
            ws.cell(row=row_num, column=3, value=supplier.contact_person)
            ws.cell(row=row_num, column=4, value=supplier.email)
            ws.cell(row=row_num, column=5, value=supplier.phone)
            ws.cell(row=row_num, column=6, value=supplier.mobile)
            ws.cell(row=row_num, column=7, value=supplier.city)
            ws.cell(row=row_num, column=8, value=supplier.country)
            ws.cell(row=row_num, column=9, value=supplier.get_payment_term_display())
            ws.cell(row=row_num, column=10, value=supplier.rating or '')
            ws.cell(row=row_num, column=11, value=float(supplier.get_balance()))
            ws.cell(row=row_num, column=12, value='Oui' if supplier.is_active else 'Non')
            ws.cell(row=row_num, column=13, value=supplier.created_at.strftime('%d/%m/%Y'))
        
        ExcelExporter.auto_adjust_columns(ws)
        
        filename = f"fournisseurs_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExcelExporter.generate_response(wb, filename)

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import suppliers from Excel file."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Aucun fichier fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
    
        try:
            df = pd.read_excel(file)
            
            # Required columns
            required_columns = ['Nom']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Colonnes manquantes: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_count = 0
            updated_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    name = str(row.get('Nom', '')).strip()
                    if not name:
                        errors.append(f"Ligne {index + 2}: Nom obligatoire")
                        continue
                    
                    email = str(row.get('Email', '')).strip() if pd.notna(row.get('Email')) else ''
                    phone = str(row.get('Téléphone', '')).strip() if pd.notna(row.get('Téléphone')) else ''
                    supplier_code = str(row.get('Code Fournisseur', '')).strip() if pd.notna(row.get('Code Fournisseur')) else ''

                    # Prepare supplier data
                    supplier_data = {
                        'name': name,
                        'contact_person': str(row.get('Contact Principal', '')).strip() if pd.notna(row.get('Contact Principal')) else '',
                        'email': email,
                        'phone': phone,
                        'mobile': str(row.get('Mobile', '')).strip() if pd.notna(row.get('Mobile')) else '',
                        'website': str(row.get('Site Web', '')).strip() if pd.notna(row.get('Site Web')) else '',
                        'address': str(row.get('Adresse', '')).strip() if pd.notna(row.get('Adresse')) else '',
                        'city': str(row.get('Ville', '')).strip() if pd.notna(row.get('Ville')) else '',
                        'postal_code': str(row.get('Code Postal', '')).strip() if pd.notna(row.get('Code Postal')) else '',
                        'country': str(row.get('Pays', 'Cameroun')).strip() if pd.notna(row.get('Pays')) else 'Cameroun',
                        'tax_id': str(row.get('Numéro Fiscal', '')).strip() if pd.notna(row.get('Numéro Fiscal')) else '',
                        'bank_account': str(row.get('Compte Bancaire', '')).strip() if pd.notna(row.get('Compte Bancaire')) else '',
                        'notes': str(row.get('Notes', '')).strip() if pd.notna(row.get('Notes')) else '',
                    }
                    
                    # Handle rating
                    if pd.notna(row.get('Évaluation')):
                        try:
                            rating = int(row.get('Évaluation'))
                            if 1 <= rating <= 5:
                                supplier_data['rating'] = rating
                        except:
                            pass
                    
                    # Handle payment term
                    payment_term_map = {
                        'Comptant': 'immediate',
                        '15 jours': '15_days',
                        '30 jours': '30_days',
                        '60 jours': '60_days',
                        '90 jours': '90_days',
                    }
                    payment_term_str = str(row.get('Conditions Paiement', '30 jours')).strip()
                    supplier_data['payment_term'] = payment_term_map.get(payment_term_str, '30_days')

                    # Create or update supplier
                    if supplier_code and Supplier.objects.filter(supplier_code=supplier_code).exists():
                        # Update existing supplier
                        supplier = Supplier.objects.get(supplier_code=supplier_code)
                        for key, value in supplier_data.items():
                            setattr(supplier, key, value)
                        supplier.save()
                        updated_count += 1
                    else:
                        # Create new supplier
                        if not supplier_code:
                            # Auto-generate code
                            count = Supplier.objects.count() + created_count + 1
                            supplier_code = f"FRN{count:05d}"
                        
                        supplier_data['supplier_code'] = supplier_code
                        Supplier.objects.create(**supplier_data)
                        created_count += 1

                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            return Response({
                'message': 'Import terminé',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'import: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Export suppliers to PDF with filters support."""
        # Use filtered queryset
        suppliers = self.filter_queryset(self.get_queryset())
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph("Liste des Fournisseurs", title_style)
        elements.append(title)
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        date_text = Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            date_style
        )
        elements.append(date_text)
        elements.append(Spacer(1, 20))
        
        # Données tableau
        data = [['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Actif']]
        for supplier in suppliers:
            row = [
                supplier.supplier_code,
                supplier.name,
                supplier.email,
                supplier.phone,
                supplier.city,
                'Oui' if supplier.is_active else 'Non'
            ]
            data.append(row)
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(table)
        
        # Stats
        elements.append(Spacer(1, 20))
        stats_text = f"Total: {suppliers.count()} fournisseur(s) | Actifs: {suppliers.filter(is_active=True).count()}"
        stats = Paragraph(stats_text, date_style)
        elements.append(stats)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f'fournisseurs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class SupplierPaymentViewSet(viewsets.ModelViewSet):
    """Manage supplier payments (create/list/retrieve)."""
    queryset = SupplierPayment.objects.all()
    serializer_class = SupplierPaymentSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'head', 'options']
    filterset_fields = ['supplier', 'payment_method', 'payment_date']
    ordering_fields = ['payment_date', 'amount', 'created_at']
    ordering = ['payment_date']
    pagination_class = None  # Désactiver la pagination pour avoir tous les paiements

    def get_queryset(self):
        """Return payments for current tenant with optional filters."""
        queryset = SupplierPayment.objects.select_related(
            'supplier', 'purchase_order', 'created_by'
        ).all()
        
        user = self.request.user
        
        # Super admin voit tout
        if user.is_superuser:
            pass  # Retourner tout le queryset
        elif hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                pass  # Retourner tout le queryset
            elif user.role.access_scope == 'own':
                queryset = queryset.filter(created_by=user)
        else:
            # Par défaut, filtrer par created_by
            queryset = queryset.filter(created_by=user)
        
        # Filter by supplier if provided
        supplier_id = self.request.query_params.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        
        return queryset

    def perform_create(self, serializer):
        """Create payment and attach current user as creator."""
        serializer.save(created_by=self.request.user)

