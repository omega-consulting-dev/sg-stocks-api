from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q, Sum, F
from apps.cashbox.models import Cashbox, CashboxSession, CashMovement
from apps.cashbox.serializers import (
    CashboxSerializer, 
    CashboxSessionSerializer, 
    CashMovementSerializer,
    EncaissementSerializer
)
from apps.invoicing.models import InvoicePayment
from apps.sales.models import Sale
from apps.expenses.models import Expense
from apps.suppliers.models import SupplierPayment
from apps.loans.models import LoanPayment
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class CashboxViewSet(viewsets.ModelViewSet):
    queryset = Cashbox.objects.select_related('store')
    serializer_class = CashboxSerializer
    filterset_fields = ['store', 'is_active']


class CashboxSessionViewSet(viewsets.ModelViewSet):
    queryset = CashboxSession.objects.select_related('cashbox', 'cashier')
    serializer_class = CashboxSessionSerializer
    filterset_fields = ['cashbox', 'cashier', 'status']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                return queryset.filter(cashbox__store__in=user.assigned_stores.all())
            elif user.role.access_scope == 'own':
                return queryset.filter(cashier=user)
        
        return queryset.filter(cashier=user)
    
    @action(detail=False, methods=['post'])
    def open_session(self, request):
        """Open a new cashbox session."""
        cashbox_id = request.data.get('cashbox')
        opening_balance = request.data.get('opening_balance', 0)
        
        # Check if there's already an open session
        if CashboxSession.objects.filter(cashbox_id=cashbox_id, status='open').exists():
            return Response(
                {'error': 'Une session est dÃ©jÃ  ouverte pour cette caisse.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        session = CashboxSession.objects.create(
            cashbox_id=cashbox_id,
            cashier=request.user,
            opening_date=timezone.now(),
            opening_balance=opening_balance,
            status='open',
            created_by=request.user
        )
        
        serializer = self.get_serializer(session)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def close_session(self, request, pk=None):
        """Close a cashbox session."""
        session = self.get_object()
        
        if session.status != 'open':
            return Response(
                {'error': 'Cette session est dÃ©jÃ  fermÃ©e.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        actual_balance = request.data.get('actual_closing_balance')
        closing_notes = request.data.get('closing_notes', '')
        
        # Calculate expected balance
        from django.db.models import Sum, Case, When, F
        movements = CashMovement.objects.filter(cashbox_session=session)
        total_in = movements.filter(movement_type='in').aggregate(total=Sum('amount'))['total'] or 0
        total_out = movements.filter(movement_type='out').aggregate(total=Sum('amount'))['total'] or 0
        
        session.expected_closing_balance = session.opening_balance + total_in - total_out
        session.actual_closing_balance = actual_balance
        session.closing_date = timezone.now()
        session.closing_notes = closing_notes
        session.status = 'closed'
        session.save()
        
        serializer = self.get_serializer(session)
        return Response(serializer.data)


class CashMovementViewSet(viewsets.ModelViewSet):
    queryset = CashMovement.objects.select_related('cashbox_session', 'sale')
    serializer_class = CashMovementSerializer
    filterset_fields = ['cashbox_session', 'movement_type', 'category', 'payment_method']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                return queryset.filter(cashbox_session__cashbox__store__in=user.assigned_stores.all())
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        return queryset.filter(created_by=user)
    
    def create(self, request, *args, **kwargs):
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Received data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.error(f"Validation errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_create(self, serializer):
        # Validation: Pour les sorties d'argent, vÃ©rifier que la caisse a assez de fonds
        if serializer.validated_data.get('movement_type') == 'out':
            cashbox_session = serializer.validated_data.get('cashbox_session')
            amount = serializer.validated_data.get('amount', 0)
            
            if cashbox_session and cashbox_session.cashbox:
                cashbox = cashbox_session.cashbox
                
                # Calculer le solde rÃ©el basÃ© sur les transactions
                from apps.cashbox.utils import get_cashbox_real_balance
                available_balance = get_cashbox_real_balance(store_id=cashbox.store.id)
                
                if amount > available_balance:
                    from rest_framework.exceptions import ValidationError
                    raise ValidationError({
                        'amount': f'Solde insuffisant. Solde disponible: {available_balance:,.2f} FCFA, Montant demandÃ©: {amount:,.2f} FCFA'
                    })
        
        # Generate movement number
        last_movement = CashMovement.objects.order_by('-id').first()
        if last_movement and last_movement.movement_number:
            try:
                last_number = int(last_movement.movement_number.replace('MVT', ''))
                next_number = last_number + 1
            except (ValueError, AttributeError):
                next_number = CashMovement.objects.count() + 1
        else:
            next_number = 1
        
        movement_number = f"MVT{next_number:08d}"
        while CashMovement.objects.filter(movement_number=movement_number).exists():
            next_number += 1
            movement_number = f"MVT{next_number:08d}"
        
        movement = serializer.save(
            movement_number=movement_number,
            created_by=self.request.user
        )


class EncaissementsListView(APIView):
    """
    Vue pour lister tous les encaissements (paiements de factures + ventes payÃ©es)
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        store_id = request.query_params.get('store')
        
        # Validation des dates
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start > end:
                    return Response(
                        {'error': 'La date de dÃ©but ne peut pas Ãªtre supÃ©rieure Ã  la date de fin'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        encaissements = []
        
        # 1. RÃ©cupÃ©rer les paiements de factures
        invoice_payments = InvoicePayment.objects.select_related('invoice', 'invoice__customer', 'invoice__store').all()
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass  # Voir tout
                elif user.role.access_scope == 'assigned':
                    invoice_payments = invoice_payments.filter(invoice__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    invoice_payments = invoice_payments.filter(invoice__created_by=user)
            else:
                invoice_payments = invoice_payments.filter(invoice__created_by=user)
        
        if start_date:
            invoice_payments = invoice_payments.filter(payment_date__gte=start_date)
        if end_date:
            invoice_payments = invoice_payments.filter(payment_date__lte=end_date)
        if store_id:
            invoice_payments = invoice_payments.filter(invoice__store_id=store_id)
        
        # Filtrer uniquement les paiements en cash (pas carte ni virement)
        invoice_payments = invoice_payments.filter(payment_method='cash')
        
        for payment in invoice_payments:
            encaissements.append({
                'id': f'INV-{payment.id}',
                'code': payment.payment_number,
                'type': 'invoice_payment',
                'date': payment.payment_date,
                'reference_facture': payment.invoice.invoice_number,
                'montant': float(payment.amount),
                'mode_paiement': payment.get_payment_method_display(),
                'client': payment.invoice.customer.name if payment.invoice.customer else '',
                'created_at': payment.created_at,
            })
        
        # 2. RÃ©cupÃ©rer les ventes avec paiement
        sales = Sale.objects.select_related('customer', 'store').filter(paid_amount__gt=0)
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass  # Voir tout
                elif user.role.access_scope == 'assigned':
                    sales = sales.filter(store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    sales = sales.filter(created_by=user)
            else:
                sales = sales.filter(created_by=user)
        
        if start_date:
            sales = sales.filter(sale_date__gte=start_date)
        if end_date:
            sales = sales.filter(sale_date__lte=end_date)
        if store_id:
            sales = sales.filter(store_id=store_id)
        
        # Filtrer uniquement les ventes payÃ©es en cash
        sales = sales.filter(payment_method='cash')
        
        for sale in sales:
            encaissements.append({
                'id': f'SALE-{sale.id}',
                'code': sale.sale_number,
                'type': 'sale',
                'date': sale.sale_date,
                'reference_facture': sale.sale_number,
                'montant': float(sale.paid_amount),
                'mode_paiement': 'EspÃ¨ces',
                'client': sale.customer.name if sale.customer else 'Client anonyme',
                'created_at': sale.created_at,
            })
        
        # Trier par date dÃ©croissante
        encaissements.sort(key=lambda x: x['date'], reverse=True)
        
        # Paginer les rÃ©sultats
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        paginated = encaissements[start_idx:end_idx]
        
        return Response({
            'count': len(encaissements),
            'next': None if end_idx >= len(encaissements) else f'?page={page + 1}',
            'previous': None if page == 1 else f'?page={page - 1}',
            'results': paginated
        })


class EncaissementsExportView(APIView):
    """
    Vue pour exporter les encaissements en Excel
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        store_id = request.query_params.get('store')
        
        # Validation des dates
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start > end:
                    return Response(
                        {'error': 'La date de dÃ©but ne peut pas Ãªtre supÃ©rieure Ã  la date de fin'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        encaissements = []
        
        # 1. RÃ©cupÃ©rer les paiements de factures
        invoice_payments = InvoicePayment.objects.select_related('invoice', 'invoice__customer', 'invoice__store').all()
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass
                elif user.role.access_scope == 'assigned':
                    invoice_payments = invoice_payments.filter(invoice__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    invoice_payments = invoice_payments.filter(invoice__created_by=user)
            else:
                invoice_payments = invoice_payments.filter(invoice__created_by=user)
        
        if start_date:
            invoice_payments = invoice_payments.filter(payment_date__gte=start_date)
        if end_date:
            invoice_payments = invoice_payments.filter(payment_date__lte=end_date)
        if store_id:
            invoice_payments = invoice_payments.filter(invoice__store_id=store_id)
        
        # Filtrer uniquement les paiements en cash (pas carte ni virement)
        invoice_payments = invoice_payments.filter(payment_method='cash')
        
        for payment in invoice_payments:
            encaissements.append({
                'code': payment.payment_number,
                'type': 'Paiement Facture',
                'date': payment.payment_date,
                'reference_facture': payment.invoice.invoice_number,
                'montant': float(payment.amount),
                'mode_paiement': payment.get_payment_method_display(),
                'client': payment.invoice.customer.name if payment.invoice.customer else '',
            })
        
        # 2. RÃ©cupÃ©rer les ventes avec paiement
        sales = Sale.objects.select_related('customer', 'store').filter(paid_amount__gt=0)
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass
                elif user.role.access_scope == 'assigned':
                    sales = sales.filter(store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    sales = sales.filter(created_by=user)
            else:
                sales = sales.filter(created_by=user)
        
        if start_date:
            sales = sales.filter(sale_date__gte=start_date)
        if end_date:
            sales = sales.filter(sale_date__lte=end_date)
        if store_id:
            sales = sales.filter(store_id=store_id)
        
        # Filtrer uniquement les ventes payÃ©es en cash
        sales = sales.filter(payment_method='cash')
        
        for sale in sales:
            encaissements.append({
                'code': sale.sale_number,
                'type': 'Vente',
                'date': sale.sale_date,
                'reference_facture': sale.sale_number,
                'montant': float(sale.paid_amount),
                'mode_paiement': 'Vente directe',
                'client': sale.customer.name if sale.customer else 'Client anonyme',
            })
        
        # Trier par date
        encaissements.sort(key=lambda x: x['date'])
        
        # CrÃ©er le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Encaissements"
        
        # Style pour l'en-tÃªte
        header_fill = PatternFill(start_color="5932EA", end_color="5932EA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tÃªtes
        headers = ['Code', 'Type', 'Date', 'RÃ©fÃ©rence', 'Montant (FCFA)', 'Mode de paiement', 'Client']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # DonnÃ©es
        for row_num, enc in enumerate(encaissements, 2):
            ws.cell(row=row_num, column=1, value=enc['code'])
            ws.cell(row=row_num, column=2, value=enc['type'])
            ws.cell(row=row_num, column=3, value=enc['date'].strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=4, value=enc['reference_facture'])
            ws.cell(row=row_num, column=5, value=enc['montant'])
            ws.cell(row=row_num, column=6, value=enc['mode_paiement'])
            ws.cell(row=row_num, column=7, value=enc['client'])
        
        # Ajouter une ligne de total
        total_row = len(encaissements) + 2
        ws.cell(row=total_row, column=4, value="TOTAL")
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        total_montant = sum(enc['montant'] for enc in encaissements)
        ws.cell(row=total_row, column=5, value=total_montant)
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # PrÃ©parer la rÃ©ponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nom du fichier avec pÃ©riode si applicable
        filename = 'encaissements'
        if start_date and end_date:
            filename += f'_{start_date}_au_{end_date}'
        elif start_date:
            filename += f'_depuis_{start_date}'
        elif end_date:
            filename += f'_jusquau_{end_date}'
        filename += '.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response


class CaisseSoldeView(APIView):
    """
    Vue pour calculer le solde actuel de la caisse basÃ© sur les transactions rÃ©elles.
    Utilise la fonction utilitaire get_cashbox_real_balance pour garantir la cohÃ©rence.
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer le paramÃ¨tre store
        store_id = request.query_params.get('store')
        
        # Filtrage par utilisateur pour les non-admins
        if not (user.is_superuser or (hasattr(user, 'role') and user.role and user.role.access_scope == 'all')):
            # Utilisateur normal : filtrer par stores assignÃ©s
            if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                # Si un store spÃ©cifique est demandÃ©, vÃ©rifier qu'il est assignÃ©
                if store_id:
                    if not user.assigned_stores.filter(id=store_id).exists():
                        return Response({'solde_actuel': 0.0})
                else:
                    # Calculer le total pour tous les stores assignÃ©s
                    from apps.cashbox.utils import get_cashbox_real_balance
                    total_balance = sum(
                        get_cashbox_real_balance(store_id=s.id) 
                        for s in user.assigned_stores.all()
                    )
                    return Response({'solde_actuel': float(total_balance)})
            else:
                return Response({'solde_actuel': 0.0})
        
        # Admin ou access_scope='all' : calculer le solde
        from apps.cashbox.utils import get_cashbox_real_balance
        
        if store_id:
            # Calculer pour un store spÃ©cifique
            real_balance = get_cashbox_real_balance(store_id=store_id)
        else:
            # Calculer le total pour tous les stores
            from apps.inventory.models import Store
            all_stores = Store.objects.filter(is_active=True)
            real_balance = sum(
                get_cashbox_real_balance(store_id=s.id) 
                for s in all_stores
            )
        
        return Response({
            'solde_actuel': float(real_balance),
        })


class DecaissementsListView(APIView):
    """
    Vue pour lister tous les dÃ©caissements (approvisionnements bancaires)
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        store_id = request.query_params.get('store')
        
        # Validation des dates
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start > end:
                    return Response(
                        {'error': 'La date de dÃ©but ne peut pas Ãªtre supÃ©rieure Ã  la date de fin'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        decaissements = []
        
        # RÃ©cupÃ©rer tous les mouvements de caisse de type "out" avec catÃ©gorie "bank_deposit"
        cash_movements = CashMovement.objects.filter(
            movement_type='out',
            category='bank_deposit'
        ).select_related('cashbox_session', 'cashbox_session__cashbox', 'cashbox_session__cashbox__store').order_by('created_at')
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass  # Voir tout
                elif user.role.access_scope == 'assigned':
                    cash_movements = cash_movements.filter(cashbox_session__cashbox__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    cash_movements = cash_movements.filter(created_by=user)
            else:
                cash_movements = cash_movements.filter(created_by=user)
        
        if start_date:
            cash_movements = cash_movements.filter(created_at__date__gte=start_date)
        if end_date:
            cash_movements = cash_movements.filter(created_at__date__lte=end_date)
        if store_id:
            cash_movements = cash_movements.filter(cashbox_session__cashbox__store_id=store_id)
        
        for movement in cash_movements:
            decaissements.append({
                'id': movement.id,  # ID numÃ©rique du CashMovement
                'code': movement.movement_number,
                'type': 'Approvisionnement Bancaire',
                'date': movement.created_at.date(),
                'reference': movement.reference or movement.movement_number,
                'montant': float(movement.amount),
                'mode_paiement': movement.get_payment_method_display(),
                'payment_method': movement.payment_method,  # Valeur brute pour l'Ã©dition
                'description': movement.description,
                'created_at': movement.created_at,
                'store_id': movement.cashbox_session.cashbox.store_id if movement.cashbox_session and movement.cashbox_session.cashbox else None,
            })
        
        # Trier par code (croissant)
        decaissements.sort(key=lambda x: x['code'])
        
        # Paginer les rÃ©sultats
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        paginated = decaissements[start_idx:end_idx]
        
        return Response({
            'count': len(decaissements),
            'next': None if end_idx >= len(decaissements) else f'?page={page + 1}',
            'previous': None if page == 1 else f'?page={page - 1}',
            'results': paginated
        })


class DecaissementsExportView(APIView):
    """
    Vue pour exporter les dÃ©caissements en Excel
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        store_id = request.query_params.get('store')
        
        # Validation des dates
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start > end:
                    return Response(
                        {'error': 'La date de dÃ©but ne peut pas Ãªtre supÃ©rieure Ã  la date de fin'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        decaissements = []
        
        # RÃ©cupÃ©rer tous les mouvements de caisse de type "out" avec catÃ©gorie "bank_deposit"
        cash_movements = CashMovement.objects.filter(
            movement_type='out',
            category='bank_deposit'
        ).select_related('cashbox_session', 'cashbox_session__cashbox', 'cashbox_session__cashbox__store').order_by('created_at')
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass
                elif user.role.access_scope == 'assigned':
                    cash_movements = cash_movements.filter(cashbox_session__cashbox__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    cash_movements = cash_movements.filter(created_by=user)
            else:
                cash_movements = cash_movements.filter(created_by=user)
        
        if start_date:
            cash_movements = cash_movements.filter(created_at__date__gte=start_date)
        if end_date:
            cash_movements = cash_movements.filter(created_at__date__lte=end_date)
        if store_id:
            cash_movements = cash_movements.filter(cashbox_session__cashbox__store_id=store_id)
        
        for movement in cash_movements:
            decaissements.append({
                'code': movement.movement_number,
                'type': 'Approvisionnement Bancaire',
                'date': movement.created_at.date(),
                'reference': movement.reference or movement.movement_number,
                'montant': float(movement.amount),
                'mode_paiement': movement.get_payment_method_display(),
                'description': movement.description,
            })
        
        # CrÃ©er le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DÃ©caissements"
        
        # Style pour l'en-tÃªte
        header_fill = PatternFill(start_color="5932EA", end_color="5932EA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tÃªtes (sans RÃ©fÃ©rence et Mode de paiement)
        headers = ['Code', 'Type', 'Date', 'Montant (FCFA)', 'Description']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # DonnÃ©es
        for row_num, dec in enumerate(decaissements, 2):
            ws.cell(row=row_num, column=1, value=dec['code'])
            ws.cell(row=row_num, column=2, value=dec['type'])
            ws.cell(row=row_num, column=3, value=dec['date'].strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=4, value=dec['montant'])
            ws.cell(row=row_num, column=5, value=dec['description'])
        
        # Ajouter une ligne de total
        total_row = len(decaissements) + 2
        ws.cell(row=total_row, column=3, value="TOTAL")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        total_montant = sum(dec['montant'] for dec in decaissements)
        ws.cell(row=total_row, column=4, value=total_montant)
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # PrÃ©parer la rÃ©ponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nom du fichier avec pÃ©riode si applicable
        filename = 'decaissements'
        if start_date and end_date:
            filename += f'_{start_date}_au_{end_date}'
        elif start_date:
            filename += f'_depuis_{start_date}'
        elif end_date:
            filename += f'_jusquau_{end_date}'
        filename += '.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response


class DecaissementsExportPDFView(APIView):
    """
    Vue pour exporter les dÃ©caissements en PDF
    """
    
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from io import BytesIO
        
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        store_id = request.query_params.get('store')
        
        # Validation des dates
        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d').date()
                end = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start > end:
                    return Response(
                        {'error': 'La date de dÃ©but ne peut pas Ãªtre supÃ©rieure Ã  la date de fin'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except ValueError:
                return Response(
                    {'error': 'Format de date invalide. Utilisez YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        decaissements = []
        
        # RÃ©cupÃ©rer tous les mouvements de caisse de type "out" avec catÃ©gorie "bank_deposit"
        cash_movements = CashMovement.objects.filter(
            movement_type='out',
            category='bank_deposit'
        ).select_related('cashbox_session', 'cashbox_session__cashbox', 'cashbox_session__cashbox__store').order_by('created_at')
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass
                elif user.role.access_scope == 'assigned':
                    cash_movements = cash_movements.filter(cashbox_session__cashbox__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    cash_movements = cash_movements.filter(created_by=user)
            else:
                cash_movements = cash_movements.filter(created_by=user)
        
        if start_date:
            cash_movements = cash_movements.filter(created_at__date__gte=start_date)
        if end_date:
            cash_movements = cash_movements.filter(created_at__date__lte=end_date)
        if store_id:
            cash_movements = cash_movements.filter(cashbox_session__cashbox__store_id=store_id)
        
        for movement in cash_movements:
            decaissements.append({
                'code': movement.movement_number,
                'type': 'Approvisionnement Bancaire',
                'date': movement.created_at.date(),
                'montant': float(movement.amount),
                'description': movement.description,
                'store': movement.cashbox_session.cashbox.store.name if movement.cashbox_session and movement.cashbox_session.cashbox else 'N/A',
            })
        
        # CrÃ©er le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=30, bottomMargin=30)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=5,
            fontName='Helvetica-Bold',
            alignment=1
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#6b7280'),
            spaceAfter=20,
            alignment=1
        )
        
        # En-tÃªte
        period = ''
        if start_date and end_date:
            period = f"PÃ©riode: {datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')} - {datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        elif start_date:
            period = f"Depuis le {datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        elif end_date:
            period = f"Jusqu'au {datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        
        elements.append(Paragraph("Liste des DÃ©caissements", title_style))
        if period:
            elements.append(Paragraph(period, subtitle_style))
        elements.append(Spacer(1, 20))
        
        # Tableau des donnÃ©es
        data = [['Code', 'Type', 'Date', 'Point de Vente', 'Montant (FCFA)', 'Description']]
        
        for dec in decaissements:
            data.append([
                dec['code'],
                dec['type'],
                dec['date'].strftime('%d/%m/%Y'),
                dec['store'],
                f"{dec['montant']:,.0f}",
                dec['description'][:50] + '...' if len(dec['description']) > 50 else dec['description']
            ])
        
        # Ligne de total
        total_montant = sum(dec['montant'] for dec in decaissements)
        data.append(['', '', '', 'TOTAL', f"{total_montant:,.0f}", ''])
        
        # CrÃ©er le tableau
        table = Table(data, colWidths=[80, 140, 70, 100, 90, 240])
        table.setStyle(TableStyle([
            # En-tÃªte
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5932EA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Corps du tableau
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#1a1a1a')),
            ('ALIGN', (0, 1), (0, -2), 'LEFT'),
            ('ALIGN', (1, 1), (3, -2), 'LEFT'),
            ('ALIGN', (4, 1), (4, -2), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -2), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TOPPADDING', (0, 1), (-1, -2), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
            
            # Ligne de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1a1a1a')),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#5932EA')),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Pied de page
        footer_text = Paragraph(
            f"<para align='center'><font size='8' color='#9ca3af'>Document gÃ©nÃ©rÃ© le {datetime.now().strftime('%d/%m/%Y Ã  %H:%M')} - Total: {len(decaissements)} dÃ©caissement(s)</font></para>",
            styles['Normal']
        )
        elements.append(footer_text)
        
        # Construire le PDF
        doc.build(elements)
        buffer.seek(0)
        
        # PrÃ©parer la rÃ©ponse
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        
        # Nom du fichier
        filename = 'decaissements'
        if start_date and end_date:
            filename += f'_{start_date}_au_{end_date}'
        elif start_date:
            filename += f'_depuis_{start_date}'
        elif end_date:
            filename += f'_jusquau_{end_date}'
        filename += '.pdf'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class StoreListView(APIView):
    """
    Vue simple pour rÃ©cupÃ©rer la liste des stores actifs
    """
    
    def get(self, request):
        from apps.inventory.models import Store
        
        stores = Store.objects.filter(is_active=True).values('id', 'name', 'code', 'store_type')
        return Response(list(stores))


class BankTransactionsListView(APIView):
    """
    Vue pour lister toutes les transactions bancaires (dÃ©pÃ´ts et retraits)
    """
    
    def get(self, request):
        user = request.user
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.GET.get('date_debut') or request.GET.get('start_date')
        end_date = request.GET.get('date_fin') or request.GET.get('end_date')
        transaction_type = request.GET.get('type')  # 'depot' ou 'retrait'
        
        # Normaliser transaction_type : None si vide ou None
        if not transaction_type or transaction_type == '':
            transaction_type = None
        
        # Valider et parser les dates
        from datetime import datetime
        parsed_start_date = None
        parsed_end_date = None
        
        if start_date:
            try:
                parsed_start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if end_date:
            try:
                parsed_end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # RÃ©cupÃ©rer les dÃ©pÃ´ts bancaires (category='bank_deposit')
        # Anciens dÃ©pÃ´ts: movement_type='out' (sortie de caisse vers banque)
        # Nouveaux dÃ©pÃ´ts: movement_type='in' (entrÃ©e dans la banque)
        deposits = CashMovement.objects.filter(
            category='bank_deposit'
        ).select_related('cashbox_session', 'cashbox_session__cashbox', 'cashbox_session__cashbox__store')
        
        # RÃ©cupÃ©rer les retraits bancaires (category='bank_withdrawal')
        # Anciens retraits: movement_type='in' (entrÃ©e dans la caisse depuis la banque)
        # Nouveaux retraits: movement_type='out' (sortie de la banque)
        withdrawals = CashMovement.objects.filter(
            category='bank_withdrawal'
        ).select_related('cashbox_session', 'cashbox_session__cashbox', 'cashbox_session__cashbox__store')
        
        # RÃ©cupÃ©rer les paiements par virement bancaire (Expenses, SupplierPayments, LoanPayments)
        from apps.expenses.models import Expense
        from apps.suppliers.models import SupplierPayment
        from apps.loans.models import LoanPayment
        
        expenses = Expense.objects.filter(
            status='paid',
            payment_method='bank_transfer'
        ).select_related('store')
        
        supplier_payments = SupplierPayment.objects.filter(
            payment_method='bank_transfer'
        ).select_related('purchase_order__store', 'supplier')
        
        loan_payments = LoanPayment.objects.filter(
            payment_method='bank_transfer'
        ).select_related('loan__store', 'loan')
        
        # RÃ©cupÃ©rer les paiements de factures par carte et virement bancaire
        from apps.invoicing.models import InvoicePayment
        
        invoice_payments = InvoicePayment.objects.filter(
            payment_method__in=['card', 'transfer'],
            status='success'
        ).select_related('invoice', 'invoice__customer', 'invoice__store')
        
        # Filtrage selon le rÃ´le
        if not user.is_superuser:
            if hasattr(user, 'role') and user.role:
                if user.role.access_scope == 'all':
                    pass
                elif user.role.access_scope == 'assigned':
                    # Filtrer par magasin assignÃ©, en gÃ©rant les dÃ©pÃ´ts sans session (cashbox_session=None)
                    from django.db.models import Q
                    deposits = deposits.filter(
                        Q(cashbox_session__isnull=True) |
                        Q(cashbox_session__cashbox__store__in=user.assigned_stores.all())
                    )
                    withdrawals = withdrawals.filter(cashbox_session__cashbox__store__in=user.assigned_stores.all())
                    expenses = expenses.filter(store__in=user.assigned_stores.all())
                    supplier_payments = supplier_payments.filter(purchase_order__store__in=user.assigned_stores.all())
                    loan_payments = loan_payments.filter(loan__store__in=user.assigned_stores.all())
                    invoice_payments = invoice_payments.filter(invoice__store__in=user.assigned_stores.all())
                elif user.role.access_scope == 'own':
                    deposits = deposits.filter(created_by=user)
                    withdrawals = withdrawals.filter(created_by=user)
                    expenses = expenses.filter(created_by=user)
                    supplier_payments = supplier_payments.filter(created_by=user)
                    loan_payments = loan_payments.filter(created_by=user)
                    invoice_payments = invoice_payments.filter(created_by=user)
            else:
                deposits = deposits.filter(created_by=user)
                withdrawals = withdrawals.filter(created_by=user)
                expenses = expenses.filter(created_by=user)
                supplier_payments = supplier_payments.filter(created_by=user)
                loan_payments = loan_payments.filter(created_by=user)
                invoice_payments = invoice_payments.filter(created_by=user)
        
        # Filtrage par date
        if parsed_start_date:
            deposits = deposits.filter(created_at__date__gte=parsed_start_date)
            withdrawals = withdrawals.filter(created_at__date__gte=parsed_start_date)
            expenses = expenses.filter(payment_date__gte=parsed_start_date)
            supplier_payments = supplier_payments.filter(payment_date__gte=parsed_start_date)
            loan_payments = loan_payments.filter(payment_date__gte=parsed_start_date)
            invoice_payments = invoice_payments.filter(payment_date__gte=parsed_start_date)
        if parsed_end_date:
            deposits = deposits.filter(created_at__date__lte=parsed_end_date)
            withdrawals = withdrawals.filter(created_at__date__lte=parsed_end_date)
            expenses = expenses.filter(payment_date__lte=parsed_end_date)
            supplier_payments = supplier_payments.filter(payment_date__lte=parsed_end_date)
            loan_payments = loan_payments.filter(payment_date__lte=parsed_end_date)
            invoice_payments = invoice_payments.filter(payment_date__lte=parsed_end_date)
        
        transactions = []
        
        # Ajouter les dÃ©pÃ´ts seulement si on filtre par 'depot' ou si aucun filtre n'est appliquÃ©
        if transaction_type is None or transaction_type == 'depot':
            for movement in deposits:
                store_name = 'N/A'
                if movement.cashbox_session and movement.cashbox_session.cashbox and movement.cashbox_session.cashbox.store:
                    store_name = movement.cashbox_session.cashbox.store.name
                
                transactions.append({
                    'id': f"dep-{movement.id}",
                    'date': movement.created_at.isoformat(),
                    'type': 'depot',
                    'amount': float(movement.amount),
                    'description': movement.description or 'DÃ©pÃ´t bancaire',
                    'store_name': store_name,
                    'balance_after': 0  # Sera calculÃ© aprÃ¨s le tri
                })
        
            # Ajouter les paiements de factures par carte et virement bancaire (entrÃ©es)
            for payment in invoice_payments:
                payment_method_display = 'Carte' if payment.payment_method == 'card' else 'Virement bancaire'
                transactions.append({
                    'id': f"inv-{payment.id}",
                    'date': payment.created_at.isoformat(),
                    'type': 'depot',
                    'amount': float(payment.amount),
                    'description': f"Paiement facture {payment.invoice.invoice_number} - {payment_method_display}",
                    'store_name': payment.invoice.store.name if payment.invoice and payment.invoice.store else 'N/A',
                    'balance_after': 0
                })
        
        # Ajouter les retraits seulement si on filtre par 'retrait' ou si aucun filtre n'est appliquÃ©
        if transaction_type is None or transaction_type == 'retrait':
            for movement in withdrawals:
                store_name = 'N/A'
                if movement.cashbox_session and movement.cashbox_session.cashbox and movement.cashbox_session.cashbox.store:
                    store_name = movement.cashbox_session.cashbox.store.name
                
                transactions.append({
                    'id': f"wit-{movement.id}",
                    'date': movement.created_at.isoformat(),
                    'type': 'retrait',
                    'amount': float(movement.amount),
                    'description': movement.description or 'Retrait bancaire',
                    'store_name': store_name,
                    'balance_after': 0  # Sera calculÃ© aprÃ¨s le tri
                })
            
            # Ajouter les dÃ©penses payÃ©es par virement bancaire
            for expense in expenses:
                # Utiliser created_at pour avoir l'heure prÃ©cise
                transaction_datetime = expense.created_at
                transactions.append({
                    'id': f"exp-{expense.id}",
                    'date': transaction_datetime.isoformat(),
                    'type': 'retrait',
                    'amount': float(expense.amount),
                    'description': f"DÃ©pense {expense.expense_number} - {expense.description or expense.category.name}",
                    'store_name': expense.store.name if expense.store else 'N/A',
                    'balance_after': 0
                })
            
            # Ajouter les paiements fournisseurs par virement bancaire
            for payment in supplier_payments:
                # Utiliser created_at pour avoir l'heure prÃ©cise
                transaction_datetime = payment.created_at
                transactions.append({
                    'id': f"sup-{payment.id}",
                    'date': transaction_datetime.isoformat(),
                    'type': 'retrait',
                    'amount': float(payment.amount),
                    'description': f"RÃ¨glement fournisseur {payment.supplier.name} par virement bancaire",
                    'store_name': payment.purchase_order.store.name if payment.purchase_order and payment.purchase_order.store else 'N/A',
                    'balance_after': 0
                })
            
            # Ajouter les remboursements d'emprunts par virement bancaire
            for payment in loan_payments:
                # Utiliser created_at pour avoir l'heure prÃ©cise
                transaction_datetime = payment.created_at
                transactions.append({
                    'id': f"loan-{payment.id}",
                    'date': transaction_datetime.isoformat(),
                    'type': 'retrait',
                    'amount': float(payment.amount),
                    'description': f"Remboursement emprunt {payment.loan.loan_number} par virement bancaire",
                    'store_name': payment.loan.store.name if payment.loan and payment.loan.store else 'N/A',
                    'balance_after': 0
                })
        
        # Trier par date
        transactions.sort(key=lambda x: x['date'])
        
        # Calculer les soldes
        balance = 0
        for transaction in transactions:
            if transaction['type'] == 'depot':
                balance += transaction['amount']
            else:  # retrait
                balance -= transaction['amount']
            transaction['balance_after'] = balance
        
        # Les transactions sont dÃ©jÃ  triÃ©es par ordre croissant de date
        
        # Calculer les totaux
        total_deposits = sum(t['amount'] for t in transactions if t['type'] == 'depot')
        total_withdrawals = sum(t['amount'] for t in transactions if t['type'] == 'retrait')
        
        # Paginer
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        paginated = transactions[start_idx:end_idx]
        
        return Response({
            'count': len(transactions),
            'next': None if end_idx >= len(transactions) else f'?page={page + 1}',
            'previous': None if page == 1 else f'?page={page - 1}',
            'results': paginated,
            'balance': balance,
            'total_deposits': total_deposits,
            'total_withdrawals': total_withdrawals
        })


class BankWithdrawalCreateView(APIView):
    """
    Vue pour crÃ©er un retrait bancaire
    """
    
    def post(self, request):
        amount = request.data.get('amount')
        description = request.data.get('description', 'Retrait bancaire')
        date = request.data.get('date')
        
        if not amount:
            return Response(
                {'error': 'Le montant est requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # GÃ©nÃ©rer le numÃ©ro de mouvement
        last_movement = CashMovement.objects.filter(
            movement_number__startswith='BWD-'
        ).order_by('-created_at').first()
        
        if last_movement:
            last_num = int(last_movement.movement_number.split('-')[1])
            movement_number = f'BWD-{last_num + 1:05d}'
        else:
            movement_number = 'BWD-00001'
        
        # CrÃ©er le mouvement de retrait bancaire (sans session de caisse)
        movement = CashMovement.objects.create(
            movement_number=movement_number,
            cashbox_session=None,  # Pas de session pour les retraits bancaires
            movement_type='out',  # Argent qui sort de la banque
            category='bank_withdrawal',
            amount=amount,
            payment_method='bank_transfer',
            description=description,
            created_by=request.user
        )
        
        # Mettre Ã  jour created_at si une date personnalisÃ©e est fournie
        if date:
            try:
                from datetime import datetime
                parsed_date = datetime.strptime(date, '%Y-%m-%d')
                movement_date = timezone.make_aware(parsed_date)
                movement.created_at = movement_date
                movement.save(update_fields=['created_at'])
            except ValueError:
                pass
        
        return Response({
            'id': movement.id,
            'date': movement.created_at.isoformat(),
            'type': 'retrait',
            'amount': float(movement.amount),
            'description': movement.description,
            'movement_number': movement.movement_number
        }, status=status.HTTP_201_CREATED)


class BankDepositCreateView(APIView):
    """
    Vue pour crÃ©er un dÃ©pÃ´t bancaire
    """
    
    def post(self, request):
        amount = request.data.get('amount')
        description = request.data.get('description', 'DÃ©pÃ´t bancaire')
        date = request.data.get('date')
        
        if not amount:
            return Response(
                {'error': 'Le montant est requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = float(amount)
            if amount <= 0:
                return Response(
                    {'error': 'Le montant doit Ãªtre supÃ©rieur Ã  0'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Montant invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # GÃ©nÃ©rer le numÃ©ro de mouvement
        last_movement = CashMovement.objects.filter(
            movement_number__startswith='BDP-'
        ).order_by('-created_at').first()
        
        if last_movement:
            last_num = int(last_movement.movement_number.split('-')[1])
            movement_number = f'BDP-{last_num + 1:05d}'
        else:
            movement_number = 'BDP-00001'
        
        # GÃ©rer la date personnalisÃ©e
        movement_date = timezone.now()
        if date:
            try:
                from datetime import datetime
                parsed_date = datetime.strptime(date, '%Y-%m-%d')
                movement_date = timezone.make_aware(parsed_date)
            except ValueError:
                pass
        
        # CrÃ©er le mouvement de dÃ©pÃ´t bancaire (sans session de caisse)
        movement = CashMovement.objects.create(
            movement_number=movement_number,
            cashbox_session=None,  # Pas de session pour les dÃ©pÃ´ts bancaires
            movement_type='in',  # Argent qui entre dans la banque
            category='bank_deposit',
            amount=amount,
            payment_method='bank_transfer',
            description=description,
            created_by=request.user
        )
        
        # Mettre Ã  jour created_at si une date personnalisÃ©e est fournie
        if date:
            movement.created_at = movement_date
            movement.save(update_fields=['created_at'])
        
        return Response({
            'id': movement.id,
            'date': movement.created_at.isoformat(),
            'type': 'depot',
            'amount': float(movement.amount),
            'description': movement.description,
            'store_name': '',
            'movement_number': movement.movement_number
        }, status=status.HTTP_201_CREATED)


class BankTransactionsExportPDFView(APIView):
    """
    Vue pour exporter les transactions bancaires en PDF
    """
    
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.GET.get('date_debut') or request.GET.get('start_date')
        end_date = request.GET.get('date_fin') or request.GET.get('end_date')
        transaction_type = request.GET.get('type')
        
        # RÃ©utiliser la logique de BankTransactionsListView
        view = BankTransactionsListView()
        view.request = request
        response_data = view.get(request).data
        
        transactions = response_data['results']
        balance = response_data['balance']
        total_deposits = response_data['total_deposits']
        total_withdrawals = response_data['total_withdrawals']
        
        # CrÃ©er le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Titre
        title = Paragraph("HISTORIQUE DES TRANSACTIONS BANCAIRES", title_style)
        elements.append(title)
        
        # PÃ©riode
        if start_date and end_date:
            period = Paragraph(
                f"PÃ©riode: du {start_date} au {end_date}",
                styles['Normal']
            )
            elements.append(period)
            elements.append(Spacer(1, 12))
        
        # RÃ©sumÃ©
        summary_data = [
            ['Solde Bancaire', f'{balance:,.2f} FCFA'],
            ['Total DÃ©pÃ´ts', f'{total_deposits:,.2f} FCFA'],
            ['Total Retraits', f'{total_withdrawals:,.2f} FCFA']
        ]
        summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des transactions
        table_data = [['Date', 'Type', 'Description', 'Magasin', 'Montant', 'Solde AprÃ¨s']]
        
        # Style pour les paragraphes dans le tableau
        desc_style = ParagraphStyle(
            'DescStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        
        for transaction in transactions:
            date_obj = datetime.fromisoformat(transaction['date'].replace('Z', '+00:00'))
            formatted_date = date_obj.strftime('%d/%m/%Y %H:%M')
            transaction_type_display = 'DÃ©pÃ´t' if transaction['type'] == 'depot' else 'Retrait'
            amount_display = f"+{transaction['amount']:,.2f}" if transaction['type'] == 'depot' else f"-{transaction['amount']:,.2f}"
            
            # Utiliser Paragraph pour la description afin de permettre le retour Ã  la ligne
            description_para = Paragraph(transaction['description'], desc_style)
            
            table_data.append([
                formatted_date,
                transaction_type_display,
                description_para,  # Utiliser le Paragraph au lieu du texte brut
                transaction['store_name'],
                amount_display,
                f"{transaction['balance_after']:,.2f}"
            ])
        
        table = Table(table_data, colWidths=[3.5*cm, 2.5*cm, 7*cm, 4*cm, 3*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alignement vertical en haut
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
        ]))
        elements.append(table)
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        
        filename = 'transactions_bancaires'
        if start_date and end_date:
            filename += f'_{start_date}_au_{end_date}'
        filename += '.pdf'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class BankTransactionsExportExcelView(APIView):
    """
    Vue pour exporter les transactions bancaires en Excel
    """
    
    def get(self, request):
        # RÃ©cupÃ©rer les paramÃ¨tres de filtre
        start_date = request.GET.get('date_debut') or request.GET.get('start_date')
        end_date = request.GET.get('date_fin') or request.GET.get('end_date')
        transaction_type = request.GET.get('type')
        
        # RÃ©utiliser la logique de BankTransactionsListView
        view = BankTransactionsListView()
        view.request = request
        response_data = view.get(request).data
        
        transactions = response_data['results']
        balance = response_data['balance']
        total_deposits = response_data['total_deposits']
        total_withdrawals = response_data['total_withdrawals']
        
        # CrÃ©er le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Bancaires"
        
        # Style du titre
        title_font = Font(size=14, bold=True, color='1e40af')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        
        # Titre
        ws['A1'] = 'HISTORIQUE DES TRANSACTIONS BANCAIRES'
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # PÃ©riode
        if start_date and end_date:
            ws['A2'] = f'PÃ©riode: du {start_date} au {end_date}'
            ws.merge_cells('A2:F2')
        
        # RÃ©sumÃ©
        ws['A4'] = 'Solde Bancaire:'
        ws['B4'] = f'{balance:,.2f} FCFA'
        ws['A5'] = 'Total DÃ©pÃ´ts:'
        ws['B5'] = f'{total_deposits:,.2f} FCFA'
        ws['A6'] = 'Total Retraits:'
        ws['B6'] = f'{total_withdrawals:,.2f} FCFA'
        
        for row in range(4, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        # En-tÃªtes du tableau
        headers = ['Date', 'Type', 'Description', 'Magasin', 'Montant', 'Solde AprÃ¨s']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # DonnÃ©es
        for row_idx, transaction in enumerate(transactions, 9):
            date_obj = datetime.fromisoformat(transaction['date'].replace('Z', '+00:00'))
            formatted_date = date_obj.strftime('%d/%m/%Y %H:%M')
            transaction_type_display = 'DÃ©pÃ´t' if transaction['type'] == 'depot' else 'Retrait'
            amount_display = f"+{transaction['amount']:,.2f}" if transaction['type'] == 'depot' else f"-{transaction['amount']:,.2f}"
            
            ws.cell(row=row_idx, column=1).value = formatted_date
            ws.cell(row=row_idx, column=2).value = transaction_type_display
            ws.cell(row=row_idx, column=3).value = transaction['description']
            ws.cell(row=row_idx, column=4).value = transaction['store_name']
            ws.cell(row=row_idx, column=5).value = amount_display
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row_idx, column=6).value = f"{transaction['balance_after']:,.2f}"
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='right')
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Sauvegarder dans un buffer
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = 'transactions_bancaires'
        if start_date and end_date:
            filename += f'_{start_date}_au_{end_date}'
        filename += '.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

class MobileMoneyBalanceView(APIView):
    """
    Vue pour recuperer le solde actuel du Mobile Money (MTN/Orange).
    """
    
    def get(self, request):
        user = request.user
        store_id = request.query_params.get('store')
        
        # Filtrage par utilisateur pour les non-admins
        if not (user.is_superuser or (hasattr(user, 'role') and user.role and user.role.access_scope == 'all')):
            if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                store_ids = list(user.assigned_stores.values_list('id', flat=True))
                if store_id:
                    if int(store_id) not in store_ids:
                        return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)
                else:
                    store_id = store_ids[0] if store_ids else None
        
        from apps.cashbox.utils import get_mobile_money_balance
        balance = get_mobile_money_balance(store_id=store_id if store_id else None)
        
        return Response({'balance': balance})


class MobileMoneyTransactionsListView(APIView):
    """
    Vue pour recuperer la liste des transactions Mobile Money.
    """
    
    def get(self, request):
        user = request.user
        store_id = request.query_params.get('store')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        transaction_type = request.query_params.get('type')
        
        # Filtrage par utilisateur selon access_scope
        user_role = getattr(user, 'role', None)
        access_scope = user_role.access_scope if user_role else 'own'
        
        # Pour les non-superadmins
        if not user.is_superuser:
            if access_scope == 'all':
                # Admin peut tout voir - pas de filtre
                pass
            elif access_scope == 'assigned':
                # Filtrer par points de vente assignés
                if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                    store_ids = list(user.assigned_stores.values_list('id', flat=True))
                    if store_id:
                        if int(store_id) not in store_ids:
                            return Response({'error': 'Acces refuse'}, status=status.HTTP_403_FORBIDDEN)
                    else:
                        store_id = store_ids[0] if store_ids else None
                else:
                    return Response({'error': 'Aucun point de vente assigne'}, status=status.HTTP_403_FORBIDDEN)
            else:  # access_scope == 'own'
                # Ne voir QUE les données créées par l'utilisateur lui-même
                # Le filtrage par created_by sera appliqué plus bas
                pass
        
        # Query de base
        movements_query = Q(category__in=['mobile_money_deposit', 'mobile_money_withdrawal'])
        
        # Filtre par created_by pour access_scope='own'
        if not user.is_superuser and access_scope == 'own':
            # Pour access_scope='own', ne voir QUE les transactions créées par l'utilisateur
            # Exclure les transactions créées par d'autres utilisateurs
            movements_query &= (Q(created_by=user) | Q(created_by__isnull=True))
        
        if store_id:
            movements_query &= Q(cashbox_session__cashbox__store_id=store_id) | Q(cashbox_session__isnull=True)
        
        if start_date:
            movements_query &= Q(created_at__gte=start_date)
        if end_date:
            movements_query &= Q(created_at__lte=end_date)
        
        if transaction_type:
            if transaction_type == 'depot':
                movements_query &= Q(category='mobile_money_deposit')
            elif transaction_type == 'retrait':
                movements_query &= Q(category='mobile_money_withdrawal')
        
        movements = CashMovement.objects.filter(movements_query).order_by('-created_at')
        
        # Query pour InvoicePayment
        payments_query = Q(payment_method='mobile_money')
        
        # Filtre par created_by pour access_scope='own'
        if not user.is_superuser and access_scope == 'own':
            # Pour access_scope='own', ne voir QUE les paiements créés par l'utilisateur
            # Exclure les paiements créés par d'autres utilisateurs
            # IMPORTANT: Les paiements avec created_by=NULL sont exclus pour éviter de voir les ventes des autres
            payments_query &= Q(created_by=user)
        
        if store_id:
            payments_query &= Q(invoice__sale__store_id=store_id)
        
        if start_date:
            payments_query &= Q(created_at__gte=start_date)
        if end_date:
            payments_query &= Q(created_at__lte=end_date)
        
        if transaction_type and transaction_type in ['depot', 'retrait']:
            payments = []
        else:
            payments = InvoicePayment.objects.filter(payments_query).select_related('invoice__sale__store').order_by('-created_at')
        
        # Query pour LoanPayment (remboursements d'emprunts)
        from apps.loans.models import LoanPayment
        loan_payments_query = Q(payment_method='mobile_money')
        
        # Filtre par created_by pour access_scope='own'
        if not user.is_superuser and access_scope == 'own':
            loan_payments_query &= Q(created_by=user)
        
        if store_id:
            loan_payments_query &= Q(loan__store_id=store_id)
        
        if start_date:
            loan_payments_query &= Q(created_at__gte=start_date)
        if end_date:
            loan_payments_query &= Q(created_at__lte=end_date)
        
        if transaction_type and transaction_type in ['depot']:
            loan_payments = []
        else:
            loan_payments = LoanPayment.objects.filter(loan_payments_query).select_related('loan__store').order_by('-created_at')
        
        # Combiner les resultats
        transactions = []
        
        for movement in movements:
            transaction_type_label = 'depot' if movement.category == 'mobile_money_deposit' else 'retrait'
            transactions.append({
                'id': movement.id,
                'date': movement.created_at.isoformat(),
                'type': transaction_type_label,
                'amount': float(movement.amount),
                'description': movement.description or '',
                'store_name': '',
                'movement_number': movement.movement_number
            })
        
        for payment in payments:
            store_name = payment.invoice.sale.store.name if payment.invoice.sale.store else ''
            transactions.append({
                'id': payment.id,
                'date': payment.created_at.isoformat(),
                'type': 'paiement',
                'amount': float(payment.amount),
                'description': f"Paiement facture #{payment.invoice.invoice_number if payment.invoice.invoice_number else payment.invoice.id}",
                'store_name': store_name,
                'movement_number': payment.invoice.invoice_number if payment.invoice.invoice_number else str(payment.invoice.id)
            })
        
        for loan_payment in loan_payments:
            store_name = loan_payment.loan.store.name if loan_payment.loan.store else ''
            transactions.append({
                'id': f'loan_{loan_payment.id}',
                'date': loan_payment.created_at.isoformat(),
                'type': 'retrait',
                'amount': float(loan_payment.amount),
                'description': f"Remboursement emprunt {loan_payment.loan.loan_number}",
                'store_name': store_name,
                'movement_number': loan_payment.payment_number
            })
        
        # Query pour SupplierPayment (paiements fournisseurs)
        from apps.suppliers.models import SupplierPayment
        supplier_payments_query = Q(payment_method='mobile_money')
        
        # Filtre par created_by pour access_scope='own'
        if not user.is_superuser and access_scope == 'own':
            supplier_payments_query &= Q(created_by=user)
        
        if store_id:
            supplier_payments_query &= (Q(purchase_order__store_id=store_id) | Q(purchase_order__isnull=True))
        
        if start_date:
            supplier_payments_query &= Q(created_at__gte=start_date)
        if end_date:
            supplier_payments_query &= Q(created_at__lte=end_date)
        
        if transaction_type and transaction_type in ['depot']:
            supplier_payments = []
        else:
            supplier_payments = SupplierPayment.objects.filter(supplier_payments_query).select_related('supplier').order_by('-created_at')
        
        for supplier_payment in supplier_payments:
            transactions.append({
                'id': f'supplier_{supplier_payment.id}',
                'date': supplier_payment.created_at.isoformat(),
                'type': 'retrait',
                'amount': float(supplier_payment.amount),
                'description': f"Paiement fournisseur {supplier_payment.supplier.name if supplier_payment.supplier else 'N/A'}",
                'store_name': '',
                'movement_number': supplier_payment.payment_number if hasattr(supplier_payment, 'payment_number') else str(supplier_payment.id)
            })
        
        # Query pour Expense (dépenses)
        from apps.expenses.models import Expense
        expenses_query = Q(payment_method='mobile_money', status='paid')
        
        # Filtre par created_by pour access_scope='own'
        if not user.is_superuser and access_scope == 'own':
            expenses_query &= Q(created_by=user)
        
        if store_id:
            expenses_query &= Q(store_id=store_id)
        
        if start_date:
            expenses_query &= Q(created_at__gte=start_date)
        if end_date:
            expenses_query &= Q(created_at__lte=end_date)
        
        if transaction_type and transaction_type in ['depot']:
            expenses = []
        else:
            expenses = Expense.objects.filter(expenses_query).select_related('store').order_by('-created_at')
        
        for expense in expenses:
            store_name = expense.store.name if expense.store else ''
            transactions.append({
                'id': f'expense_{expense.id}',
                'date': expense.created_at.isoformat(),
                'type': 'retrait',
                'amount': float(expense.amount),
                'description': f"Dépense: {expense.description or expense.expense_number}",
                'store_name': store_name,
                'movement_number': expense.expense_number
            })
        
        # Trier par date croissante pour calculer le solde progressif
        transactions.sort(key=lambda x: x['date'])
        
        # Calculer le solde après chaque transaction
        from apps.cashbox.utils import get_mobile_money_balance
        from decimal import Decimal
        
        # Obtenir le solde actuel
        current_balance = get_mobile_money_balance(store_id=store_id if store_id else None)
        
        # Calculer le solde initial (avant toutes les transactions affichées)
        total_displayed = Decimal('0')
        for transaction in transactions:
            amount = Decimal(str(transaction['amount']))
            if transaction['type'] in ['depot', 'paiement']:
                total_displayed += amount
            else:  # retrait
                total_displayed -= amount
        
        initial_balance = current_balance - total_displayed
        
        # Ajouter le solde après chaque transaction
        running_balance = initial_balance
        for transaction in transactions:
            amount = Decimal(str(transaction['amount']))
            if transaction['type'] in ['depot', 'paiement']:
                running_balance += amount
            else:  # retrait
                running_balance -= amount
            transaction['balance_after'] = float(running_balance)
        
        # Garder l'ordre croissant (du plus ancien au plus récent)
        # Les transactions sont déjà triées par date croissante
        
        return Response(transactions)


class MobileMoneyDepositCreateView(APIView):
    """
    Vue pour creer un depot de la caisse vers Mobile Money.
    """
    
    def post(self, request):
        user = request.user
        amount = request.data.get('amount')
        date = request.data.get('date')
        motif = request.data.get('motif', '')
        store_id = request.query_params.get('store')
        
        # Vérification des permissions selon access_scope
        user_role = getattr(user, 'role', None)
        access_scope = user_role.access_scope if user_role else 'own'
        
        if not user.is_superuser:
            if access_scope == 'all':
                # Admin - peut créer pour n'importe quel point de vente
                pass
            elif access_scope == 'assigned':
                # Utilisateur assigné - utiliser son point de vente assigné
                if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                    store_ids = list(user.assigned_stores.values_list('id', flat=True))
                    if store_id:
                        if int(store_id) not in store_ids:
                            return Response({'error': 'Acces refuse a ce point de vente'}, status=status.HTTP_403_FORBIDDEN)
                    else:
                        store_id = store_ids[0] if store_ids else None
                else:
                    return Response({'error': 'Aucun point de vente assigne'}, status=status.HTTP_403_FORBIDDEN)
            else:  # access_scope == 'own'
                # Propres données uniquement - peut créer mais sera limité à ses propres données
                # Pas besoin de point de vente spécifique car c'est filtré par created_by
                pass
        
        if not amount:
            return Response({'error': 'Montant requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            amount = float(amount)
            if amount <= 0:
                return Response({'error': 'Le montant doit etre superieur a zero'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': 'Montant invalide'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verifier le solde de la caisse
        from apps.cashbox.utils import get_cashbox_real_balance
        cashbox_balance = get_cashbox_real_balance(store_id=store_id if store_id else None)
        
        if cashbox_balance < amount:
            return Response(
                {'error': f'Solde caisse insuffisant. Solde actuel: {cashbox_balance} FCFA'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generer le numero de mouvement
        last_movement = CashMovement.objects.filter(
            movement_number__startswith='MMDP-'
        ).order_by('-created_at').first()
        
        if last_movement:
            last_num = int(last_movement.movement_number.split('-')[1])
            movement_number = f'MMDP-{last_num + 1:05d}'
        else:
            movement_number = 'MMDP-00001'
        
        # Gerer la date personnalisee
        movement_date = timezone.now()
        if date:
            try:
                from datetime import datetime
                parsed_date = datetime.strptime(date, '%Y-%m-%d')
                movement_date = timezone.make_aware(parsed_date)
            except ValueError:
                pass
        
        # Creer le mouvement de depot Mobile Money
        movement = CashMovement.objects.create(
            movement_number=movement_number,
            cashbox_session=None,
            movement_type='out',  # Sortie de la caisse
            category='mobile_money_deposit',
            amount=amount,
            payment_method='mobile_money',
            description=motif,
            created_by=request.user
        )
        
        if date:
            movement.created_at = movement_date
            movement.save(update_fields=['created_at'])
        
        return Response({
            'id': movement.id,
            'date': movement.created_at.isoformat(),
            'type': 'depot',
            'amount': float(movement.amount),
            'description': movement.description,
            'store_name': '',
            'movement_number': movement.movement_number
        }, status=status.HTTP_201_CREATED)


class MobileMoneyWithdrawalCreateView(APIView):
    """
    Vue pour creer un retrait de Mobile Money vers la caisse.
    """
    
    def post(self, request):
        user = request.user
        amount = request.data.get('amount')
        date = request.data.get('date')
        description = request.data.get('description', '')
        store_id = request.query_params.get('store')
        
        # Vérification des permissions selon access_scope
        user_role = getattr(user, 'role', None)
        access_scope = user_role.access_scope if user_role else 'own'
        
        if not user.is_superuser:
            if access_scope == 'all':
                # Admin - peut créer pour n'importe quel point de vente
                pass
            elif access_scope == 'assigned':
                # Utilisateur assigné - utiliser son point de vente assigné
                if hasattr(user, 'assigned_stores') and user.assigned_stores.exists():
                    store_ids = list(user.assigned_stores.values_list('id', flat=True))
                    if store_id:
                        if int(store_id) not in store_ids:
                            return Response({'error': 'Acces refuse a ce point de vente'}, status=status.HTTP_403_FORBIDDEN)
                    else:
                        store_id = store_ids[0] if store_ids else None
                else:
                    return Response({'error': 'Aucun point de vente assigne'}, status=status.HTTP_403_FORBIDDEN)
            else:  # access_scope == 'own'
                # Propres données uniquement - peut créer mais sera limité à ses propres données
                pass
        
        if not amount:
            return Response({'error': 'Montant requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            amount = float(amount)
            if amount <= 0:
                return Response({'error': 'Le montant doit etre superieur a zero'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': 'Montant invalide'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verifier le solde Mobile Money
        from apps.cashbox.utils import get_mobile_money_balance
        mm_balance = get_mobile_money_balance(store_id=store_id if store_id else None)
        
        if mm_balance < amount:
            return Response(
                {'error': f'Solde Mobile Money insuffisant. Solde actuel: {mm_balance} FCFA'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generer le numero de mouvement
        last_movement = CashMovement.objects.filter(
            movement_number__startswith='MMWD-'
        ).order_by('-created_at').first()
        
        if last_movement:
            last_num = int(last_movement.movement_number.split('-')[1])
            movement_number = f'MMWD-{last_num + 1:05d}'
        else:
            movement_number = 'MMWD-00001'
        
        # Gerer la date personnalisee
        movement_date = timezone.now()
        if date:
            try:
                from datetime import datetime
                parsed_date = datetime.strptime(date, '%Y-%m-%d')
                movement_date = timezone.make_aware(parsed_date)
            except ValueError:
                pass
        
        # Creer le mouvement de retrait Mobile Money
        movement = CashMovement.objects.create(
            movement_number=movement_number,
            cashbox_session=None,
            movement_type='in',  # Entree dans la caisse
            category='mobile_money_withdrawal',
            amount=amount,
            payment_method='mobile_money',
            description=description,
            created_by=request.user
        )
        
        if date:
            movement.created_at = movement_date
            movement.save(update_fields=['created_at'])
        
        return Response({
            'id': movement.id,
            'date': movement.created_at.isoformat(),
            'type': 'retrait',
            'amount': float(movement.amount),
            'description': movement.description,
            'store_name': '',
            'movement_number': movement.movement_number
        }, status=status.HTTP_201_CREATED)


class MobileMoneyTransactionsExportPDFView(APIView):
    """
    Vue pour exporter les transactions Mobile Money en PDF
    """
    
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        from datetime import datetime
        
        # Recuperer les parametres de filtre
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        transaction_type = request.GET.get('type')
        store_id = request.GET.get('store')
        
        # Reutiliser la logique de MobileMoneyTransactionsListView
        view = MobileMoneyTransactionsListView()
        view.request = request
        transactions = view.get(request).data
        
        # Calculer le solde
        from apps.cashbox.utils import get_mobile_money_balance
        balance = get_mobile_money_balance(store_id=store_id if store_id else None)
        
        # Calculer les totaux
        total_deposits = sum(t['amount'] for t in transactions if t['type'] in ['depot', 'paiement'])
        total_withdrawals = sum(t['amount'] for t in transactions if t['type'] == 'retrait')
        
        # Creer le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#ea580c'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Titre
        title = Paragraph("HISTORIQUE DES TRANSACTIONS MOBILE MONEY (MTN/ORANGE)", title_style)
        elements.append(title)
        
        # Periode
        if start_date and end_date:
            period = Paragraph(
                f"Periode: du {start_date} au {end_date}",
                styles['Normal']
            )
            elements.append(period)
            elements.append(Spacer(1, 12))
        
        # Resume
        summary_data = [
            ['Solde Mobile Money', f'{balance:,.2f} XAF'],
            ['Total Depots', f'{total_deposits:,.2f} XAF'],
            ['Total Retraits', f'{total_withdrawals:,.2f} XAF']
        ]
        summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fed7aa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#ea580c')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des transactions
        table_data = [['Numero', 'Date', 'Type', 'Description', 'Magasin', 'Montant']]
        
        desc_style = ParagraphStyle(
            'DescStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        
        for idx, transaction in enumerate(transactions, 1):
            date_obj = datetime.fromisoformat(transaction['date'].replace('Z', '+00:00'))
            formatted_date = date_obj.strftime('%d/%m/%Y')
            transaction_type_display = 'Depot' if transaction['type'] in ['depot', 'paiement'] else 'Retrait'
            amount_display = f"+{transaction['amount']:,.2f}" if transaction['type'] in ['depot', 'paiement'] else f"-{transaction['amount']:,.2f}"
            
            description_para = Paragraph(transaction['description'], desc_style)
            
            table_data.append([
                str(idx),
                formatted_date,
                transaction_type_display,
                description_para,
                transaction['store_name'],
                amount_display
            ])
        
        table = Table(table_data, colWidths=[2*cm, 3*cm, 3*cm, 7*cm, 5*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ea580c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="transactions_mobile_money.pdf"'
        return response


class MobileMoneyTransactionsExportExcelView(APIView):
    """
    Vue pour exporter les transactions Mobile Money en Excel
    """
    
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from datetime import datetime
        
        # Recuperer les parametres de filtre
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        transaction_type = request.GET.get('type')
        store_id = request.GET.get('store')
        
        # Reutiliser la logique de MobileMoneyTransactionsListView
        view = MobileMoneyTransactionsListView()
        view.request = request
        transactions = view.get(request).data
        
        # Calculer le solde
        from apps.cashbox.utils import get_mobile_money_balance
        balance = get_mobile_money_balance(store_id=store_id if store_id else None)
        
        # Calculer les totaux
        total_deposits = sum(t['amount'] for t in transactions if t['type'] in ['depot', 'paiement'])
        total_withdrawals = sum(t['amount'] for t in transactions if t['type'] == 'retrait')
        
        # Creer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions Mobile Money"
        
        # Style pour l'en-tete
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'] = "HISTORIQUE DES TRANSACTIONS MOBILE MONEY (MTN/ORANGE)"
        ws['A1'].font = Font(bold=True, size=14, color="EA580C")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Resume
        ws['A3'] = "Solde Mobile Money:"
        ws['B3'] = f"{balance:,.2f} XAF"
        ws['A4'] = "Total Depots:"
        ws['B4'] = f"{total_deposits:,.2f} XAF"
        ws['A5'] = "Total Retraits:"
        ws['B5'] = f"{total_withdrawals:,.2f} XAF"
        
        # En-tetes du tableau
        headers = ['Numero', 'Date', 'Type', 'Description', 'Magasin', 'Montant']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Donnees
        for idx, transaction in enumerate(transactions, 1):
            row = idx + 7
            date_obj = datetime.fromisoformat(transaction['date'].replace('Z', '+00:00'))
            formatted_date = date_obj.strftime('%d/%m/%Y')
            transaction_type_display = 'Depot' if transaction['type'] in ['depot', 'paiement'] else 'Retrait'
            amount = transaction['amount'] if transaction['type'] in ['depot', 'paiement'] else -transaction['amount']
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=formatted_date)
            ws.cell(row=row, column=3, value=transaction_type_display)
            ws.cell(row=row, column=4, value=transaction['description'])
            ws.cell(row=row, column=5, value=transaction['store_name'])
            ws.cell(row=row, column=6, value=amount)
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transactions_mobile_money.xlsx"'
        return response
