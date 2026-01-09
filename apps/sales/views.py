from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import HasModulePermission
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view
from django.db.models import Sum, Count
from django.utils import timezone

from core.utils.export_utils import ExcelExporter, PDFExporter

from apps.sales.models import Sale, Quote, SaleLine
from apps.sales.serializers import (
    SaleListSerializer, SaleDetailSerializer, SaleCreateSerializer,
    QuoteSerializer
)


@extend_schema_view(
    list=extend_schema(summary="Liste des ventes", tags=["Sales"]),
    retrieve=extend_schema(summary="Détail d'une vente", tags=["Sales"]),
    create=extend_schema(summary="Créer une vente", tags=["Sales"]),
)
class SaleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Sale model with secure user-based filtering.
    - Super admin: voit toutes les ventes
    - Manager (access_scope='all'): voit toutes les ventes
    - Caissier (access_scope='assigned'): voit les ventes de ses stores assignés
    - Caissier (access_scope='own'): voit uniquement ses propres ventes
    """
    
    queryset = Sale.objects.select_related('customer', 'store').prefetch_related('lines')
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'sales'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['customer', 'store', 'status', 'payment_status', 'sale_date']
    search_fields = ['sale_number', 'customer__username']
    ordering_fields = ['sale_date', 'total_amount', 'created_at']
    ordering = ['sale_date']
    
    def get_queryset(self):
        """
        Filtrage sécurisé des ventes selon le rôle et access_scope.
        Chaque caissier ne voit que ses propres ventes.
        """
        queryset = super().get_queryset()
        user = self.request.user
        
        # Super admin voit tout
        if user.is_superuser:
            return queryset
        
        # Vérifier le scope d'accès du rôle
        if hasattr(user, 'role') and user.role:
            # Manager avec accès à toutes les ventes
            if user.role.access_scope == 'all':
                return queryset
            
            # Utilisateur avec accès aux stores assignés
            elif user.role.access_scope == 'assigned':
                assigned_stores = user.assigned_stores.all()
                if assigned_stores.exists():
                    return queryset.filter(store__in=assigned_stores)
                else:
                    return queryset.none()
            
            # Utilisateur avec accès uniquement à ses propres ventes (caissiers)
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        # Par défaut, filtrer par créateur (sécurité)
        return queryset.filter(created_by=user)
    
    def get_serializer_class(self):
        if self.action == 'list':
            return SaleListSerializer
        elif self.action == 'create' or self.action == 'update':
            return SaleCreateSerializer
        return SaleDetailSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def create(self, request, *args, **kwargs):
        """Override create to return detail serializer in response."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        
        # Use detail serializer for response
        instance = serializer.instance
        detail_serializer = SaleDetailSerializer(instance)
        headers = self.get_success_headers(detail_serializer.data)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    @extend_schema(summary="Confirmer une vente et décrémenter le stock", tags=["Sales"])
    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def confirm(self, request, pk=None):
        """Confirm sale and decrement stock."""
        # Vérifier explicitement la permission can_manage_sales
        if not request.user.is_superuser:
            if not hasattr(request.user, 'role') or not request.user.role:
                return Response(
                    {'error': 'Vous n\'avez pas les permissions nécessaires'},
                    status=status.HTTP_403_FORBIDDEN
                )
            if not request.user.role.can_manage_sales:
                return Response(
                    {'error': 'Vous n\'avez pas les permissions pour confirmer des ventes'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        sale = self.get_object()
        
        try:
            sale.confirm()
            # Reload the sale instance to get the invoice created by the signal
            sale.refresh_from_db()
            
            return Response({
                'message': 'Vente confirmée avec succès',
                'sale_id': sale.id,
                'sale_number': sale.sale_number,
                'status': sale.status
            })
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @extend_schema(summary="Générer une facture depuis une vente confirmée", tags=["Sales"])
    @action(detail=True, methods=['post'])
    def generate_invoice(self, request, pk=None):
        """Generate invoice from confirmed sale."""
        sale = self.get_object()
        
        try:
            from apps.invoicing.models import Invoice
            from apps.invoicing.serializers import InvoiceDetailSerializer
            
            invoice = Invoice.generate_from_sale(sale)
            invoice_serializer = InvoiceDetailSerializer(invoice)
            
            return Response({
                'message': 'Facture générée avec succès',
                'invoice': invoice_serializer.data
            })
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @extend_schema(summary="Marquer une vente comme complétée", tags=["Sales"])
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark sale as completed."""
        sale = self.get_object()
        
        try:
            sale.complete()
            serializer = self.get_serializer(sale)
            return Response({
                'message': 'Vente complétée avec succès',
                'sale': serializer.data
            })
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @extend_schema(summary="Annuler une vente et restaurer le stock", tags=["Sales"])
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel sale and restore stock if confirmed."""
        sale = self.get_object()
        
        try:
            sale.cancel()
            serializer = self.get_serializer(sale)
            return Response({
                'message': 'Vente annulée avec succès',
                'sale': serializer.data
            })
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @extend_schema(summary="Statistiques des ventes par produit/catégorie", tags=["Sales"])
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get sales statistics grouped by product, service or category."""
        from apps.products.models import Product, ProductCategory
        import datetime
        
        # Get filters
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        product_id = request.query_params.get('product')
        category_id = request.query_params.get('category')
        store_id = request.query_params.get('store')
        group_by = request.query_params.get('group_by', 'product')  # 'product', 'service' or 'category'
        line_type = request.query_params.get('line_type')  # Optional: 'product', 'service', or both if not specified
        
        # Base queryset - only confirmed and completed sales
        sale_lines = SaleLine.objects.filter(
            sale__status__in=['confirmed', 'completed']
        ).select_related('product', 'product__category', 'service', 'sale', 'sale__store')
        
        # Filter by line type if specified
        if line_type == 'product':
            sale_lines = sale_lines.filter(line_type='product', product__isnull=False)
        elif line_type == 'service':
            sale_lines = sale_lines.filter(line_type='service', service__isnull=False)
        # If line_type not specified, include both products and services
        
        # Apply date filters
        if date_from:
            try:
                date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
                sale_lines = sale_lines.filter(sale__sale_date__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
                sale_lines = sale_lines.filter(sale__sale_date__lte=date_to_obj)
            except ValueError:
                pass
        
        # Apply store filter
        if store_id:
            sale_lines = sale_lines.filter(sale__store_id=store_id)
        
        result = []
        
        if category_id:
            # Group by category
            category = ProductCategory.objects.filter(id=category_id).first()
            if category:
                category_lines = sale_lines.filter(product__category=category)
                total_ca = sum(
                    float(line.quantity) * float(line.unit_price) 
                    for line in category_lines
                )
                total_qty = sum(float(line.quantity) for line in category_lines)
                
                # Only add if there are actual sales
                if total_ca > 0:
                    result.append({
                        'reference': f'CAT-{category.id:03d}',
                        'designation': category.name,
                        'ca': total_ca,
                        'quantity': total_qty
                    })
        elif product_id:
            # Filter by specific product
            product = Product.objects.filter(id=product_id).first()
            if product:
                product_lines = sale_lines.filter(product=product)
                total_ca = sum(
                    float(line.quantity) * float(line.unit_price) 
                    for line in product_lines
                )
                total_qty = sum(float(line.quantity) for line in product_lines)
                
                # Only add if there are actual sales
                if total_ca > 0:
                    result.append({
                        'reference': product.reference or f'PROD-{product.id:03d}',
                        'designation': product.name,
                        'ca': total_ca,
                        'quantity': total_qty
                    })
        else:
            # Group by product, service or category based on group_by parameter
            if group_by == 'category':
                # Group by all categories (products only)
                categories_data = {}
                for line in sale_lines:
                    if line.line_type == 'product' and line.product and line.product.category:
                        cat_id = line.product.category.id
                        if cat_id not in categories_data:
                            categories_data[cat_id] = {
                                'category': line.product.category,
                                'ca': 0,
                                'quantity': 0
                            }
                        categories_data[cat_id]['ca'] += float(line.quantity) * float(line.unit_price)
                        categories_data[cat_id]['quantity'] += float(line.quantity)
                
                # Convert to list and only include categories with sales
                for cat_id, data in categories_data.items():
                    if data['ca'] > 0:
                        result.append({
                            'reference': f'CAT-{cat_id:03d}',
                            'designation': data['category'].name,
                            'ca': data['ca'],
                            'quantity': data['quantity']
                        })
            elif group_by == 'service':
                # Group by service
                services_data = {}
                for line in sale_lines:
                    if line.line_type == 'service' and line.service:
                        serv_id = line.service.id
                        if serv_id not in services_data:
                            services_data[serv_id] = {
                                'service': line.service,
                                'ca': 0,
                                'quantity': 0
                            }
                        services_data[serv_id]['ca'] += float(line.quantity) * float(line.unit_price)
                        services_data[serv_id]['quantity'] += float(line.quantity)
                
                # Only add services that have sales
                for serv_id, data in services_data.items():
                    if data['ca'] > 0:
                        result.append({
                            'reference': data['service'].reference or f'SERV-{serv_id:03d}',
                            'designation': data['service'].name,
                            'ca': data['ca'],
                            'quantity': data['quantity']
                        })
            else:
                # Group by product or service - both
                items_data = {}
                for line in sale_lines:
                    if line.line_type == 'product' and line.product:
                        item_key = f'P-{line.product.id}'
                        if item_key not in items_data:
                            items_data[item_key] = {
                                'type': 'product',
                                'item': line.product,
                                'ca': 0,
                                'quantity': 0
                            }
                        items_data[item_key]['ca'] += float(line.quantity) * float(line.unit_price)
                        items_data[item_key]['quantity'] += float(line.quantity)
                    elif line.line_type == 'service' and line.service:
                        item_key = f'S-{line.service.id}'
                        if item_key not in items_data:
                            items_data[item_key] = {
                                'type': 'service',
                                'item': line.service,
                                'ca': 0,
                                'quantity': 0
                            }
                        items_data[item_key]['ca'] += float(line.quantity) * float(line.unit_price)
                        items_data[item_key]['quantity'] += float(line.quantity)
                
                # Add items that have sales
                for item_key, data in items_data.items():
                    if data['ca'] > 0:
                        if data['type'] == 'product':
                            result.append({
                                'reference': data['item'].reference or f'PROD-{data["item"].id:03d}',
                                'designation': data['item'].name,
                                'ca': data['ca'],
                                'quantity': data['quantity'],
                                'type': 'product'
                            })
                        else:  # service
                            result.append({
                                'reference': data['item'].reference or f'SERV-{data["item"].id:03d}',
                                'designation': data['item'].name,
                                'ca': data['ca'],
                                'quantity': data['quantity'],
                                'type': 'service'
                            })
            
            # Sort by CA descending
            result.sort(key=lambda x: x['ca'], reverse=True)
        
        return Response(result)
    
    @extend_schema(summary="Exporter les ventes en Excel", tags=["Sales"])
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export sales to Excel."""
        sales = self.filter_queryset(self.get_queryset())
        
        wb, ws = ExcelExporter.create_workbook("Ventes")
        
        columns = [
            'N° Vente', 'Date', 'Client', 'Magasin', 'Montant Total',
            'Montant Payé', 'Statut', 'Statut Paiement'
        ]
        ExcelExporter.style_header(ws, columns)
        
        for row_num, sale in enumerate(sales, 2):
            ws.cell(row=row_num, column=1, value=sale.sale_number)
            ws.cell(row=row_num, column=2, value=sale.sale_date.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=3, value=sale.customer.name if sale.customer else 'N/A')
            ws.cell(row=row_num, column=4, value=sale.store.name)
            ws.cell(row=row_num, column=5, value=float(sale.total_amount))
            ws.cell(row=row_num, column=6, value=float(sale.paid_amount))
            ws.cell(row=row_num, column=7, value=sale.get_status_display())
            ws.cell(row=row_num, column=8, value=sale.get_payment_status_display())
        
        ExcelExporter.auto_adjust_columns(ws)
        
        filename = f"ventes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ExcelExporter.generate_response(wb, filename)


    @extend_schema(summary="Exporter les ventes en PDF", tags=["Sales"])
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Export sales to PDF."""
        sales = self.filter_queryset(self.get_queryset())[:100]
        
        buffer = io.BytesIO()
        doc = PDFExporter.create_document(buffer)
        styles = PDFExporter.get_styles()
        story = []
        
        story.append(Paragraph("Rapport des Ventes", styles['CustomTitle']))
        story.append(Spacer(1, 0.5*inch))
        
        # Summary
        total_amount = sum(sale.total_amount for sale in sales)
        story.append(Paragraph(f"Total: {total_amount:,.0f} XAF", styles['CustomSubtitle']))
        story.append(Spacer(1, 0.3*inch))
        
        # Table
        data = [['N° Vente', 'Date', 'Client', 'Montant', 'Statut']]
        for sale in sales:
            data.append([
                sale.sale_number,
                sale.sale_date.strftime('%d/%m/%Y'),
                sale.customer.name[:20] if sale.customer else 'N/A',
                f"{sale.total_amount:,.0f}",
                sale.get_status_display()
            ])
        
        table = PDFExporter.create_table(data)
        story.append(table)
        
        doc.build(story)
        
        filename = f"ventes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return PDFExporter.generate_response(buffer, filename)
    
    @extend_schema(summary="Exporter statistiques en Excel", tags=["Sales"])
    @action(detail=False, methods=['get'], url_path='export_statistics_excel')
    def export_statistics_excel(self, request):
        """Export sales statistics to Excel."""
        # Get statistics data
        statistics_response = self.stats(request)
        statistics_data = statistics_response.data
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques Ventes"
        
        columns = ['Réf.', 'Désignation', 'C. A.']
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, col_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_num, stat in enumerate(statistics_data, 2):
            ws.cell(row=row_num, column=1, value=stat['reference'])
            ws.cell(row=row_num, column=2, value=stat['designation'])
            ws.cell(row=row_num, column=3, value=float(stat['ca']))
        
        # Auto adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        
        # Generate response
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = f"statistiques_ventes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    @extend_schema(summary="Exporter statistiques en PDF", tags=["Sales"])
    @action(detail=False, methods=['get'], url_path='export_statistics_pdf')
    def export_statistics_pdf(self, request):
        """Export sales statistics to PDF."""
        # Get statistics data
        statistics_response = self.stats(request)
        statistics_data = statistics_response.data
        
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#366092'),
            spaceAfter=20,
            alignment=1
        )
        
        title = Paragraph("Statistiques des Ventes", title_style)
        elements.append(title)
        
        # Date info
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        if date_from and date_to:
            period_text = f"Période : {date_from} au {date_to}"
        else:
            period_text = "Toutes les périodes"
        
        date_text = Paragraph(
            f"{period_text} - Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", 
            date_style
        )
        elements.append(date_text)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        data = [['Réf.', 'Désignation', 'C. A.']]
        
        total_ca = 0
        for stat in statistics_data:
            total_ca += float(stat['ca'])
            row = [
                stat['reference'],
                stat['designation'][:40],
                f"{stat['ca']:,.0f} FCFA"
            ]
            data.append(row)
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        
        # Statistics
        elements.append(Spacer(1, 0.3*inch))
        stats_text = f"<b>Total CA:</b> {total_ca:,.0f} FCFA | <b>Nombre d'articles:</b> {len(statistics_data)}"
        stats = Paragraph(stats_text, date_style)
        elements.append(stats)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Generate response
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f'statistiques_ventes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response



@extend_schema_view(
    list=extend_schema(summary="Liste des devis", tags=["Sales"]),
    retrieve=extend_schema(summary="Détail d'un devis", tags=["Sales"]),
    create=extend_schema(summary="Créer un devis", tags=["Sales"]),
)
class QuoteViewSet(viewsets.ModelViewSet):
    """ViewSet for Quote model."""
    
    queryset = Quote.objects.select_related('customer', 'store').prefetch_related('lines')
    serializer_class = QuoteSerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'sales'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['customer', 'store', 'status']
    search_fields = ['quote_number', 'customer__username']
    ordering_fields = ['quote_date', 'created_at']
    ordering = ['quote_date']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.is_superuser:
            return queryset
        
        if hasattr(user, 'role') and user.role:
            if user.role.access_scope == 'all':
                return queryset
            elif user.role.access_scope == 'assigned':
                return queryset.filter(store__in=user.assigned_stores.all())
            elif user.role.access_scope == 'own':
                return queryset.filter(created_by=user)
        
        return queryset.filter(created_by=user)
    
    @extend_schema(summary="Convertir en vente", tags=["Sales"])
    @action(detail=True, methods=['post'])
    def convert_to_sale(self, request, pk=None):
        """Convert quote to sale."""
        quote = self.get_object()
        
        if quote.status != 'accepted':
            return Response(
                {'error': 'Seuls les devis acceptés peuvent être convertis.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create sale from quote
        sale_data = {
            'customer': quote.customer,
            'store': quote.store,
            'sale_date': timezone.now().date(),
            'discount_amount': quote.discount_amount,
            'notes': f"Créé depuis devis {quote.quote_number}",
        }
        
        sale = Sale.objects.create(**sale_data, created_by=request.user)
        
        # Copy lines
        for quote_line in quote.lines.all():
            SaleLine.objects.create(
                sale=sale,
                line_type=quote_line.line_type,
                product=quote_line.product,
                service=quote_line.service,
                description=quote_line.description,
                quantity=quote_line.quantity,
                unit_price=quote_line.unit_price,
                tax_rate=quote_line.tax_rate,
                discount_percentage=quote_line.discount_percentage,
            )
        
        sale.calculate_totals()
        sale.save()
        
        # Link quote to sale
        quote.sale = sale
        quote.save()
        
        from apps.sales.serializers import SaleDetailSerializer
        serializer = SaleDetailSerializer(sale)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
